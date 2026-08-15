"""Rey Capital — combined trade dashboard for the ExecRelay dev stack.

Single-file stdlib HTTP server, branded with the Rey Capital design system
(colors/typography/components mirrored from C:\\AccountManagementSystem
frontend/src/index.css). Combines, on every request:

  * transactions/mt5-fills.log*         -- every order the EA shim executed
    (all TradingView-sourced; "other" covers non-stack activity on the
    account, e.g. manual trades or a different EA)
  * the running MT5 terminal (optional) -- open positions + a selectable
    trailing window (7/30/90d) of closed deals for the shim's magic number,
    with realized P/L (partial closes grouped into one row per position)
  * .local-stack/journal.json           -- lightweight trading journal keyed
    by MT5 position ticket; fields mirror the ReyLens journal schema
    (setup / emotion / mistakes / rating / notes / reviewed) so entries can
    be migrated into ReyLens later.

Started by scripts/local-stack.ps1 as service "trade-dashboard". Binds to
localhost only. Two layers of protection since the page shows account
numbers and balances:

  * Host-header check on every request (DNS-rebinding guard) -- only
    127.0.0.1:<port> / localhost:<port> (port derived from DASHBOARD_ADDR).
  * Optional bearer token (env DASHBOARD_TOKEN): when set, every route
    except /health requires it via ?token=<value> or X-Dashboard-Token.

Environment:
    DASHBOARD_ADDR   default 127.0.0.1:8090
    DASHBOARD_TOKEN  optional bearer token; unset/empty = auth disabled
    EA_SHIM_MAGIC    default 20240101 (must match ea_shim.py)
"""

from __future__ import annotations

import csv
import hmac
import io
import json
import os
import re
import sys
import threading
import time
import urllib.error
import urllib.request
from datetime import datetime, timedelta, timezone
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlsplit

try:
    import MetaTrader5 as mt5
except ImportError:  # dashboard still works log-only
    mt5 = None

import _tradestore as ts
from _txnlog import get_txn_logger, log_txn

ADDR = os.environ.get("DASHBOARD_ADDR", "127.0.0.1:8090")
MAGIC = int(os.environ.get("EA_SHIM_MAGIC", "20240101"))
DASHBOARD_TOKEN = (os.environ.get("DASHBOARD_TOKEN") or "").strip()
ROOT = Path(__file__).resolve().parent.parent

TXN_DIR = ROOT / ".local-stack" / "logs" / "transactions"
JOURNAL_PATH = ROOT / ".local-stack" / "journal.json"
ASSETS = Path(__file__).resolve().parent / "dashboard-assets"

EMOTIONS = ["calm", "confident", "neutral", "anxious", "fearful", "greedy", "fomo", "revenge", "bored"]

# ---------------------------------------------------------------------------
# ExecRelay webhook credentials -- prefixed onto every command body this
# dashboard sends straight to ingress (see "Send a sample TradingView
# signal" below). LICENSE_ID is the per-license credential ADR 0006 requires
# on every webhook command (docs/adr/0006-per-license-hmac-not-global.md);
# WEBHOOK_SECRET is the optional additional per-command secret. Sourced from
# dashboard-local env vars -- this is the only remaining thing in-repo that
# needs them locally for its own test-signal tool.
# ---------------------------------------------------------------------------

LICENSE_ID = (os.environ.get("DASHBOARD_LICENSE_ID") or "").strip()
WEBHOOK_SECRET = (os.environ.get("DASHBOARD_WEBHOOK_SECRET") or "").strip()

# ---------------------------------------------------------------------------
# Management reporting — env config
#
# Risk cap mirrors the EA shim's own cap (docs/development/windows-local-stack
# "Risk sizing — max loss per trade"); the compliance panel reports orders
# against it, it does not enforce anything itself.
# ---------------------------------------------------------------------------

EA_SHIM_RISK_USD = float(os.environ.get("EA_SHIM_RISK_USD", "0") or 0)

# ---------------------------------------------------------------------------
# Pipeline status ("is anything broken") -- see pipeline_status() below.
# Inherited from the same .env the rest of the stack sees (local-stack.ps1
# imports it into the process before spawning every service, including this
# one).
# ---------------------------------------------------------------------------


def _env_bool(name: str, default: str = "false") -> bool:
    return (os.environ.get(name, default) or "").strip().lower() in ("true", "1", "yes", "on")


_DRY_RUN_ENV_DEFAULT = _env_bool("DASHBOARD_DRY_RUN_DEFAULT", "true")
_HB_STALE_AFTER_SEC = 90.0


def effective_dry_run() -> bool:
    """The dry-run state actually in force right now.

    The operator can flip dry-run from this dashboard (POST /api/dryrun), which
    persists to the trade store's `meta` table. The env var is only the
    default for a store that has never been written to, so never branch on
    _DRY_RUN_ENV_DEFAULT directly -- always ask here."""
    return ts.get_dry_run(_DRY_RUN_ENV_DEFAULT)


def set_dry_run_mode(payload: dict) -> tuple[int, dict]:
    """Flip the dry-run switch. Returns (http_status, body).

    Gates, mirroring send_test_signal()'s posture -- the two are the only
    controls in this dashboard that can change whether real orders reach the
    broker:

      1. `enabled` must be the JSON boolean true/false, exactly. Anything
         else (missing, "false" the string, 0) is a 400 -- a malformed call
         must never be read as "go live".
      2. Turning dry-run OFF (going live) additionally requires
         `confirm: true`, so a stray click or a replayed request cannot arm
         live trading. Turning it ON is the safe direction and needs no
         confirmation.
      3. Going live also re-checks the same precondition the webhook-test
         tool relies on: LICENSE_ID must be set, because it prefixes every
         command sent to the broker. That check used to run only at boot,
         when dry-run could not change afterwards -- this switch can, so the
         guarantee has to be re-established here or the switch would be a way
         around it.
      4. Every accepted change is written to the control audit log -- going
         live is exactly the event an operator wants a receipt for.
    """
    enabled = payload.get("enabled")
    if not isinstance(enabled, bool):
        return 400, {"ok": False, "error": "enabled must be true or false"}

    if enabled is False and payload.get("confirm") is not True:
        return 400, {
            "ok": False,
            "error": "confirm must be true to turn dry-run off (this arms live trading)",
            "dry_run": effective_dry_run(),
        }

    if enabled is False and not LICENSE_ID:
        return 409, {
            "ok": False,
            "error": "DASHBOARD_LICENSE_ID is not set — it prefixes every broker "
                     "command, so going live would emit malformed orders. Set it in .env "
                     "and restart the stack before switching off dry-run.",
            "dry_run": effective_dry_run(),
        }

    before = effective_dry_run()
    ts.set_dry_run(enabled)
    after = effective_dry_run()

    log_txn(
        CONTROL_TXN_LOG,
        action="dry_run",
        before=before,
        after=after,
        requested=enabled,
        ok=(after == enabled),
    )
    return 200, {"ok": after == enabled, "dry_run": after}

# ---------------------------------------------------------------------------
# INGRESS_PORT is independently configurable (default 8081, matching
# local-stack.ps1's $IngressPort) so tests can point it at a local stub
# server without touching the real ingress. _post_to_ingress() below is the
# only place in this file that ever POSTs there -- currently used solely by
# the "Send a sample TradingView signal" tool (send_test_signal()).
# ---------------------------------------------------------------------------

INGRESS_PORT = (os.environ.get("INGRESS_PORT") or "8081").strip() or "8081"
INGRESS_PERIMETER_TOKEN = (os.environ.get("INGRESS_PERIMETER_TOKEN") or "").strip()
_INGRESS_WEBHOOK_URL = f"http://127.0.0.1:{INGRESS_PORT}/webhook"
# Operator control actions that change how the stack trades (currently just the
# dry-run switch) get their own audit trail.
CONTROL_TXN_LOG = get_txn_logger("dashboard-control")
# "Send a sample TradingView signal" (see build_test_signal_command() /
# send_test_signal() below) gets its own audit trail too.
WEBHOOK_TEST_TXN_LOG = get_txn_logger("dashboard-webhook-test")


def _allowed_hosts() -> set[str]:
    _, _, port = ADDR.rpartition(":")
    port = port or "8090"
    return {f"127.0.0.1:{port}", f"localhost:{port}"}


ALLOWED_HOSTS = _allowed_hosts()

_journal_lock = threading.Lock()

# MT5 connection cache: avoid reconnecting on every request, but don't cache
# a broken connection forever -- if account_info() ever comes back None
# (terminal closed/relaunched, logged out, etc.) we shut down and retry
# initialize() on a simple time-based backoff so a dead terminal doesn't
# turn every request into a slow retry loop.
_mt5_state = {"ready": False, "last_attempt": 0.0}
_MT5_RETRY_BACKOFF_SEC = 10.0

# Broker-server UTC offset, estimated once from a live tick and cached for
# the life of the process (see _estimate_broker_offset).
_broker_offset_state: dict = {"offset": None}

# Parsed-JSONL cache for the txn log files, keyed by path with (size, mtime)
# validation so unchanged files are not re-read/re-parsed every request.
_txn_cache_lock = threading.Lock()
_txn_cache: dict[str, tuple[int, float, list[dict]]] = {}


def _ensure_mt5() -> bool:
    if mt5 is None:
        return False
    if _mt5_state["ready"]:
        return True
    now = time.time()
    if now - _mt5_state["last_attempt"] < _MT5_RETRY_BACKOFF_SEC:
        return False
    _mt5_state["last_attempt"] = now
    _mt5_state["ready"] = bool(mt5.initialize())
    return _mt5_state["ready"]


def _mt5_reset() -> None:
    """Drop the cached "ready" connection so the next request retries
    initialize() (after the backoff) instead of trusting a stale state."""
    if mt5 is not None:
        try:
            mt5.shutdown()
        except Exception:
            pass
    _mt5_state["ready"] = False


def _estimate_broker_offset(symbol_hint: str | None) -> float | None:
    """MT5 deal/tick times are stamped in the broker server's timezone, not
    UTC. Estimate that offset from a live tick: the raw skew between the
    tick's reported time and our wall clock is (broker offset + tick age).
    Rounding the skew to the nearest 30 minutes gives the offset (brokers
    use whole/half-hour zones); if the leftover residual after rounding is
    more than ~10 minutes, the tick is too stale to trust (market closed)
    and the offset is left unknown rather than caching a bad guess."""
    if _broker_offset_state["offset"] is not None:
        return _broker_offset_state["offset"]
    if mt5 is None:
        return None
    symbol = symbol_hint or "XAUUSD_"
    try:
        tick = mt5.symbol_info_tick(symbol)
    except Exception:
        tick = None
    tick_time = getattr(tick, "time", None) if tick else None
    if not tick_time:
        return None
    delta = tick_time - time.time()
    rounded = round(delta / 1800.0) * 1800.0
    if abs(delta - rounded) > 600:
        return None
    # A stale tick (market closed over a weekend) can be hours old and still
    # round cleanly to a multiple of 30 min, masquerading as an offset. Real
    # broker offsets are a few hours (MT5 brokers cluster around UTC+2/+3),
    # so anything beyond +/-4h is staleness, not timezone -- leave unknown
    # and let times be labeled "broker time" instead of caching a bad guess.
    if abs(rounded) > 4 * 3600:
        return None
    _broker_offset_state["offset"] = rounded
    return rounded


def _fmt_time(raw: float | None, offset: float | None) -> str | None:
    if raw is None:
        return None
    ts = raw - offset if offset is not None else raw
    try:
        return datetime.fromtimestamp(ts, tz=timezone.utc).isoformat()
    except (OSError, OverflowError, ValueError):
        return None


# ---------------------------------------------------------------------------
# Data sources
# ---------------------------------------------------------------------------


def _read_one(f: Path) -> list[dict]:
    try:
        st = f.stat()
    except OSError:
        return []
    key = str(f)
    with _txn_cache_lock:
        cached = _txn_cache.get(key)
    if cached and cached[0] == st.st_size and cached[1] == st.st_mtime:
        return cached[2]
    records: list[dict] = []
    try:
        for line in f.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line:
                try:
                    records.append(json.loads(line))
                except json.JSONDecodeError:
                    pass
    except OSError:
        return []
    with _txn_cache_lock:
        _txn_cache[key] = (st.st_size, st.st_mtime, records)
    return records


def _read_txn(name: str) -> list[dict]:
    """All retained JSONL records for one txn logger, oldest first."""
    files = sorted(TXN_DIR.glob(f"{name}.log.*")) + [TXN_DIR / f"{name}.log"]
    records: list[dict] = []
    for f in files:
        records.extend(_read_one(f))
    return records


def _load_journal() -> dict:
    try:
        return json.loads(JOURNAL_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def _save_journal(journal: dict) -> None:
    tmp = JOURNAL_PATH.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(journal, indent=1), encoding="utf-8")
    tmp.replace(JOURNAL_PATH)


def _side_of(command: str) -> str:
    c = (command or "").lower()
    if c == "closeshortopenlong" or c.startswith("buy"):
        return "buy"
    if c == "closelongopenshort" or c.startswith("sell"):
        return "sell"
    return ""


def _source_of(comment: str) -> str:
    """This stack has exactly one signal source: every order it places is
    TradingView-sourced by construction. Kept as a function (rather than
    inlining the literal) so callers don't need to change if a second source
    is ever added back."""
    return "tradingview"


def _deal_source(magic: int, comment: str) -> str:
    """Classify account activity: this stack's own trades (by magic) are
    always "tradingview"; anything else on the account (other EAs, manual
    trades) is "other"."""
    if magic == MAGIC:
        return _source_of(comment)
    return "other"


def _redact(cmd: str) -> str:
    return re.sub(r"secret=[^,]*", "secret=***", cmd or "")


def _order_stats(source: str | None = None) -> dict:
    """`source`, when given, restricts to orders whose source matches
    ("tradingview"/"other"). mt5-fills.log only ever logs this stack's own
    (tradingview) orders -- "other" EA/manual activity never appears here, so
    source="other" naturally yields empty buckets/recent rows rather than
    needing special-casing."""
    rows = [r for r in _read_txn("mt5-fills") if r.get("command")]
    if source:
        rows = [r for r in rows if _source_of(r.get("comment", "")) == source]
    out = {"tradingview": _empty_bucket()}
    recent = []
    for r in rows:
        b = out[_source_of(r.get("comment", ""))]
        b["total"] += 1
        status = r.get("status")
        if status in ("filled", "placed"):
            b["executed"] += 1
        elif status == "rejected":
            b["rejected"] += 1
        side = _side_of(r.get("command", ""))
        if side == "buy":
            b["buys"] += 1
        elif side == "sell":
            b["sells"] += 1
        if r.get("event") != "position_closed":
            recent.append(
                {
                    "ts": r.get("ts", ""),
                    "source": _source_of(r.get("comment", "")),
                    "command": r.get("command", ""),
                    "symbol": r.get("symbol", ""),
                    "risk": r.get("risk") or "",
                    "volume": r.get("volume") or "",
                    "sl": r.get("sl") or "",
                    "tp": r.get("tp") or "",
                    "status": status or "",
                    "error": (r.get("error") or "")[:80],
                }
            )
    return {"by_source": out, "recent": recent[-12:][::-1]}


def _empty_bucket() -> dict:
    return {"total": 0, "executed": 0, "rejected": 0, "buys": 0, "sells": 0}


def _fetch_deals(days: int) -> list:
    if not _ensure_mt5():
        return []
    now = datetime.now()
    return list(mt5.history_deals_get(now - timedelta(days=days), now + timedelta(days=1)) or [])


def _group_closed(deals: list) -> list[dict]:
    """Group closing deals by position_id -- a position closed in several
    partial fills otherwise appears as multiple rows and double-counts
    win/loss/net. One row per position: summed profit(+commission+swap),
    summed closed volume, last close price/time; entry price (volume
    weighted) and entry time (first IN deal) pulled from the matching
    DEAL_ENTRY_IN deals in the same window when available."""
    out_entries = (mt5.DEAL_ENTRY_OUT, getattr(mt5, "DEAL_ENTRY_OUT_BY", 2))
    ins: dict[int, list] = {}
    outs: dict[int, list] = {}
    for d in deals:
        if d.type not in (mt5.DEAL_TYPE_BUY, mt5.DEAL_TYPE_SELL):
            continue
        if d.entry == mt5.DEAL_ENTRY_IN:
            ins.setdefault(d.position_id, []).append(d)
        elif d.entry in out_entries:
            outs.setdefault(d.position_id, []).append(d)

    rows: list[dict] = []
    for pos_id, out_deals in outs.items():
        out_deals = sorted(out_deals, key=lambda x: x.time)
        last = out_deals[-1]
        profit = round(sum(x.profit + x.commission + x.swap for x in out_deals), 2)
        commission = round(sum(x.commission for x in out_deals), 2)
        swap = round(sum(x.swap for x in out_deals), 2)
        volume = round(sum(x.volume for x in out_deals), 4)
        side = "sell" if last.type == mt5.DEAL_TYPE_BUY else "buy"

        in_deals = sorted(ins.get(pos_id, []), key=lambda x: x.time)
        entry_price = None
        entry_time_raw = None
        if in_deals:
            tot_vol = sum(x.volume for x in in_deals)
            if tot_vol:
                entry_price = round(sum(x.price * x.volume for x in in_deals) / tot_vol, 5)
            entry_time_raw = in_deals[0].time

        rows.append(
            {
                "ticket": str(pos_id),
                "time_raw": last.time,
                "symbol": last.symbol,
                "side": side,
                "volume": volume,
                "close": last.price,
                "entry": entry_price,
                "entry_time_raw": entry_time_raw,
                "profit": profit,
                "commission": commission,
                "swap": swap,
                "source": _deal_source(last.magic, last.comment),
                "comment": last.comment,
                "magic": last.magic,
            }
        )
    return rows


def _daily_pnl(rows: list[dict], days: int) -> list[dict]:
    buckets: dict[str, float] = {}
    for r in rows:
        d = (r.get("time") or "")[:10]
        if d:
            buckets[d] = round(buckets.get(d, 0.0) + r["profit"], 2)
    end = datetime.now(timezone.utc).date()
    start = end - timedelta(days=days - 1)
    out = []
    cur = start
    while cur <= end:
        key = cur.isoformat()
        out.append({"date": key, "pnl": buckets.get(key, 0.0)})
        cur += timedelta(days=1)
    return out


def _closed_positions(days: int, journal: dict) -> tuple[list[dict], float | None]:
    """Grouped closed positions for the trailing `days` window, newest
    first, with times converted using the estimated broker offset (see
    _estimate_broker_offset) and journal entries attached."""
    if not _ensure_mt5():
        return [], None
    positions = mt5.positions_get() or []
    offset = _estimate_broker_offset(positions[0].symbol if positions else None)
    deals = _fetch_deals(days)
    grouped = _group_closed(deals)
    for r in grouped:
        r["time"] = _fmt_time(r["time_raw"], offset)
        r["entry_time"] = _fmt_time(r["entry_time_raw"], offset) if r.get("entry_time_raw") else None
        r["journal"] = journal.get(r["ticket"]) or None
    grouped.sort(key=lambda r: r["time_raw"], reverse=True)
    return grouped, offset


def _mt5_stats(journal: dict, days: int, source: str | None = None) -> dict:
    """`source`, when given ("tradingview"/"other"), restricts open
    positions and the closed-trades window (count/wins/losses/net/rows/daily/
    review_queue) to that source. `closed.by_source` is the one exception --
    it is always computed from the *unfiltered* closed set, because the
    "Performance by source" section is a three-way comparison that must stay
    visible regardless of which source is selected (the UI highlights the
    selected row instead of hiding the others)."""
    if not _ensure_mt5():
        return {"available": False}
    acct = mt5.account_info()
    if acct is None:
        # Was ready before but the terminal stopped answering -- reset so
        # the next request retries initialize() instead of caching this
        # broken state forever.
        _mt5_reset()
        return {"available": False}

    open_rows_all = []
    positions = mt5.positions_get() or []
    for p in positions:
        open_rows_all.append(
            {
                "ticket": p.ticket,
                "symbol": p.symbol,
                "side": "buy" if p.type == mt5.POSITION_TYPE_BUY else "sell",
                "volume": p.volume,
                "entry": p.price_open,
                "sl": p.sl,
                "tp": p.tp,
                "profit": round(p.profit, 2),
                "source": _deal_source(p.magic, p.comment),
            }
        )
    open_rows = [r for r in open_rows_all if r["source"] == source] if source else open_rows_all
    open_floating = round(sum(r["profit"] for r in open_rows), 2) if open_rows else 0.0

    grouped_all, offset = _closed_positions(days, journal)
    time_label = "UTC" if offset is not None else "broker time"
    grouped = [r for r in grouped_all if r["source"] == source] if source else grouped_all

    wins = [r for r in grouped if r["profit"] >= 0]
    by_source: dict[str, float] = {"tradingview": 0.0, "other": 0.0}
    for r in grouped_all:
        by_source[r["source"]] = round(by_source.get(r["source"], 0.0) + r["profit"], 2)

    review_queue = [r for r in grouped if not r.get("journal")][:8]

    return {
        "available": True,
        "account": acct.login,
        "currency": acct.currency,
        "balance": acct.balance,
        "equity": acct.equity,
        "open": open_rows,
        "open_floating": open_floating,
        "time_label": time_label,
        "closed": {
            "days": days,
            "count": len(grouped),
            "wins": len(wins),
            "losses": len(grouped) - len(wins),
            "net": round(sum(r["profit"] for r in grouped), 2),
            "buys": sum(1 for r in grouped if r["side"] == "buy"),
            "sells": sum(1 for r in grouped if r["side"] == "sell"),
            "rows": grouped,
            "by_source": by_source,
            "daily": _daily_pnl(grouped, days),
            "review_queue": review_queue,
        },
    }


def summary(days: int = 7, source: str | None = None) -> dict:
    """`source` ("tradingview"/"other"), when given, filters orders, open
    positions and the closed-trades window throughout. The "Performance by
    source" comparison inside mt5.closed.by_source is always computed
    unfiltered -- see _mt5_stats. source=None (the default) shows everything."""
    journal = _load_journal()
    journaled = [j for j in journal.values() if j.get("setup") or j.get("notes") or j.get("rating")]
    ratings = [j["rating"] for j in journaled if j.get("rating")]
    return {
        "updated": datetime.now(timezone.utc).isoformat(),
        "days": days,
        "source": source or "",
        "orders": _order_stats(source),
        "mt5": _mt5_stats(journal, days, source),
        "journal": {
            "entries": len(journaled),
            "reviewed": sum(1 for j in journaled if j.get("reviewed")),
            "avg_rating": round(sum(ratings) / len(ratings), 1) if ratings else None,
            "emotions": EMOTIONS,
        },
        "risk": risk_panel(days, source),
    }


def save_journal_entry(payload: dict) -> dict:
    ticket = str(payload.get("ticket", "")).strip()
    if not ticket:
        raise ValueError("ticket required")
    rating = payload.get("rating")
    entry = {
        "setup": str(payload.get("setup", ""))[:120],
        "emotion": payload.get("emotion") if payload.get("emotion") in EMOTIONS else "",
        "mistakes": str(payload.get("mistakes", ""))[:500],
        "rating": int(rating) if rating and str(rating).isdigit() and 1 <= int(rating) <= 5 else None,
        "notes": str(payload.get("notes", ""))[:4000],
        "reviewed": bool(payload.get("reviewed")),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    with _journal_lock:
        journal = _load_journal()
        journal[ticket] = entry
        _save_journal(journal)
    return entry


# ---------------------------------------------------------------------------
# CSV exports
# ---------------------------------------------------------------------------

_CSV_DANGEROUS_PREFIXES = ("=", "+", "-", "@")


def _csv_cell(v) -> str:
    """CSV-injection guard: a cell opening with =, +, -, or @ can be
    interpreted as a formula by Excel/Sheets when the file is opened;
    prefix it with a single quote to force text interpretation."""
    s = "" if v is None else str(v)
    if s and s[0] in _CSV_DANGEROUS_PREFIXES:
        return "'" + s
    return s


def _write_csv(header: list[str], data_rows: list[list]) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(header)
    for row in data_rows:
        w.writerow([_csv_cell(x) for x in row])
    return buf.getvalue().encode("utf-8")


def trades_csv(days: int, source: str | None = None) -> bytes:
    """ReyLens import-template column order for closed (grouped) trades."""
    journal = _load_journal()
    rows: list[dict] = []
    if _ensure_mt5() and mt5.account_info() is not None:
        rows, _ = _closed_positions(days, journal)
        if source:
            rows = [r for r in rows if r["source"] == source]
    elif _mt5_state["ready"] and mt5 is not None and mt5.account_info() is None:
        _mt5_reset()

    data = [
        [
            "",  # account_id: ReyLens account UUID isn't known here
            r["symbol"],
            r["side"],
            r["volume"],
            r.get("entry") if r.get("entry") is not None else "",
            r["close"],
            r.get("entry_time") or "",
            r.get("time") or "",
            r["profit"],
            "",  # pnl_pct: not reliably derivable without margin/contract size
            r.get("commission", 0),
            r.get("swap", 0),
            r.get("magic", ""),
            r.get("comment", ""),
            "false",
            "closed",
        ]
        for r in rows
    ]
    header = [
        "account_id", "symbol", "side", "lot_size", "entry_price", "exit_price",
        "entry_time", "exit_time", "pnl", "pnl_pct", "commission", "swap",
        "magic_number", "comment", "is_manual", "status",
    ]
    return _write_csv(header, data)


def journal_csv() -> bytes:
    journal = _load_journal()
    lookup: dict[str, dict] = {}
    # Wide window so tickets journaled a while ago still resolve to their
    # symbol/side/close time/profit; on-demand export, not on the hot path.
    if _ensure_mt5() and mt5.account_info() is not None:
        rows, _ = _closed_positions(365, journal)
        lookup = {r["ticket"]: r for r in rows}
    elif _mt5_state["ready"] and mt5 is not None and mt5.account_info() is None:
        _mt5_reset()

    data = []
    for ticket, j in journal.items():
        r = lookup.get(ticket, {})
        data.append(
            [
                ticket,
                r.get("symbol", ""),
                r.get("side", ""),
                r.get("time", ""),
                r.get("profit", ""),
                j.get("setup", ""),
                j.get("emotion", ""),
                j.get("mistakes", ""),
                j.get("rating") if j.get("rating") is not None else "",
                j.get("reviewed", False),
                j.get("notes", ""),
                j.get("updated_at", ""),
            ]
        )
    header = [
        "ticket", "symbol", "side", "close_time", "profit", "setup", "emotion",
        "mistakes", "rating", "reviewed", "notes", "updated_at",
    ]
    return _write_csv(header, data)


def _parse_days(path: str) -> int:
    qs = parse_qs(urlsplit(path).query)
    raw = (qs.get("days") or ["7"])[0]
    try:
        d = int(raw)
    except ValueError:
        d = 7
    return max(1, min(120, d))


_VALID_SOURCES = ("tradingview", "other")


def _parse_source(path: str) -> str | None:
    qs = parse_qs(urlsplit(path).query)
    raw = (qs.get("source") or [""])[0].strip().lower()
    return raw if raw in _VALID_SOURCES else None


# ---------------------------------------------------------------------------
# Management reporting — backed by the SQLite trade store (_tradestore.py,
# .local-stack\\execrelay.db). All queries go through ts.query(), which never
# raises (returns [] on any DB error), so these helpers are safe to call even
# if the store is missing/locked/corrupt -- callers just see empty sections.
# ---------------------------------------------------------------------------


def _cutoff_iso(days: int) -> str:
    return (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()


def _channel_label(channel: str | None) -> str:
    c = (channel or "").strip()
    return c if c else "direct"


def scorecard(days: int = 30) -> dict:
    """Per-channel signal->order->closed-trade funnel. The `signals` table
    this joins against (signals.channel / signals.trace_ids -> orders.trace_id)
    is no longer populated by anything in this stack, so in practice this
    reduces to two synthetic rows: "tradingview", covering orders.source=
    "tradingview", and "other EAs", covering closed_trades.source="other" --
    together they reconcile the section's totals against the whole account,
    not just the signal-driven slice. The signals-table join is left in place
    (harmless against an empty table) rather than special-cased away, so a
    future signal source can populate it again without code changes here.

    Order -> closed-trade join is a heuristic: orders.broker_order_id ==
    closed_trades.position_id, which only resolves once MT5 reports a market
    fill's close. Pending orders (limits not yet triggered) or very recent
    fills the closed-trade backfill/live recorder hasn't caught up to show as
    "open/pending" rather than wins/losses.
    """
    cutoff = _cutoff_iso(days)
    sig_rows = ts.query("SELECT channel, outcome, trace_ids FROM signals WHERE ts >= ?", (cutoff,))
    order_rows = ts.query(
        "SELECT trace_id, source, status, requested_risk, broker_order_id FROM orders WHERE ts >= ?", (cutoff,)
    )
    orders_by_trace = {r["trace_id"]: r for r in order_rows if r["trace_id"]}
    closed_rows = ts.query("SELECT position_id, profit, source FROM closed_trades WHERE close_ts >= ?", (cutoff,))
    closed_by_pos = {r["position_id"]: r for r in closed_rows}

    channels: dict[str, dict] = {}

    def bucket(name: str) -> dict:
        return channels.setdefault(
            name,
            {
                "channel": name, "received": 0, "posted": 0, "rejected": 0,
                "orders_executed": 0, "orders_open_pending": 0,
                "wins": 0, "losses": 0, "net_pl": 0.0, "r_values": [],
            },
        )

    def _apply_order(b: dict, o: dict) -> None:
        if o.get("status") not in ("filled", "placed"):
            return
        b["orders_executed"] += 1
        pos = o.get("broker_order_id")
        closed = closed_by_pos.get(pos) if pos else None
        if closed is None:
            b["orders_open_pending"] += 1
            return
        profit = closed.get("profit") or 0.0
        if profit >= 0:
            b["wins"] += 1
        else:
            b["losses"] += 1
        b["net_pl"] += profit
        risk = o.get("requested_risk")
        if risk:
            b["r_values"].append(profit / risk)

    for s in sig_rows:
        b = bucket(_channel_label(s.get("channel")))
        b["received"] += 1
        outcome = s.get("outcome")
        if outcome == "posted":
            b["posted"] += 1
        elif outcome == "rejected":
            b["rejected"] += 1
        try:
            trace_ids = json.loads(s.get("trace_ids") or "[]")
            if not isinstance(trace_ids, list):
                trace_ids = []
        except json.JSONDecodeError:
            trace_ids = []
        for tid in trace_ids:
            o = orders_by_trace.get(tid)
            if o:
                _apply_order(b, o)

    tv = bucket("tradingview")
    for o in order_rows:
        if o.get("source") == "tradingview":
            _apply_order(tv, o)

    other = bucket("other EAs")
    for c in closed_rows:
        if c.get("source") != "other":
            continue
        profit = c.get("profit") or 0.0
        if profit >= 0:
            other["wins"] += 1
        else:
            other["losses"] += 1
        other["net_pl"] += profit

    rows = []
    for name, b in channels.items():
        avg_r = round(sum(b["r_values"]) / len(b["r_values"]), 2) if b["r_values"] else None
        rows.append(
            {
                "channel": b["channel"],
                "received": b["received"],
                "posted": b["posted"],
                "rejected": b["rejected"],
                "orders_executed": b["orders_executed"],
                "orders_open_pending": b["orders_open_pending"],
                "wins": b["wins"],
                "losses": b["losses"],
                "net_pl": round(b["net_pl"], 2),
                "avg_r": avg_r,
            }
        )
    # Real signal channels first (alphabetical), synthetic reconciliation
    # rows last in a fixed order.
    _synthetic_order = {"tradingview": 1, "other EAs": 2}
    rows.sort(key=lambda r: (_synthetic_order.get(r["channel"], 0), r["channel"]))
    return {
        "days": days,
        "rows": rows,
        "note": (
            "orders are matched to closed trades via broker_order_id == position_id "
            "(market fills only); unmatched executed orders show as open/pending "
            "until MT5 reports the close"
        ),
    }


# --- risk & exposure ---------------------------------------------------


def _equity_curve(days: int) -> dict:
    cutoff = _cutoff_iso(days)
    snaps = ts.query(
        "SELECT ts, balance, equity, margin, margin_free, floating FROM equity_snapshots "
        "WHERE ts >= ? ORDER BY ts ASC",
        (cutoff,),
    )
    estimated = len(snaps) < 2
    if not estimated:
        curve = [{"ts": s["ts"], "equity": s["equity"]} for s in snaps]
    else:
        # Fewer than 2 real snapshots in the window (e.g. ea_shim hasn't been
        # running long) -- fall back to a synthetic curve: start from the
        # current balance and reverse-cumulate each day's realized closed
        # P/L to back out what equity "would have been" on prior days. This
        # ignores floating P/L on still-open positions and any deposits/
        # withdrawals, so it's an approximation -- callers must label it.
        balance_now = None
        if _ensure_mt5():
            acct = mt5.account_info()
            if acct is not None:
                balance_now = acct.balance
        if balance_now is None and snaps:
            balance_now = snaps[-1]["balance"]
        if balance_now is None:
            balance_now = 0.0

        closed = ts.query("SELECT close_ts, profit FROM closed_trades WHERE close_ts >= ?", (cutoff,))
        daily: dict[str, float] = {}
        for c in closed:
            d = (c.get("close_ts") or "")[:10]
            if d:
                daily[d] = daily.get(d, 0.0) + (c.get("profit") or 0.0)

        end = datetime.now(timezone.utc).date()
        start = end - timedelta(days=days - 1)
        day_keys = []
        cur = start
        while cur <= end:
            day_keys.append(cur.isoformat())
            cur += timedelta(days=1)

        cum_after = 0.0
        eq_by_day: dict[str, float] = {}
        for d in reversed(day_keys):
            eq_by_day[d] = round(balance_now - cum_after, 2)
            cum_after += daily.get(d, 0.0)
        curve = [{"ts": d, "equity": eq_by_day[d]} for d in day_keys]

    max_dd = 0.0
    cur_dd = 0.0
    if curve:
        peak = curve[0]["equity"]
        for pt in curve:
            eq = pt["equity"]
            if eq > peak:
                peak = eq
            if peak:
                dd = (peak - eq) / peak * 100
                if dd > max_dd:
                    max_dd = dd
        last_eq = curve[-1]["equity"]
        cur_dd = ((peak - last_eq) / peak * 100) if peak else 0.0

    return {
        "estimated": estimated,
        "points": curve,
        "max_drawdown_pct": round(max_dd, 2),
        "current_drawdown_pct": round(cur_dd, 2),
    }


def _live_risk() -> dict:
    if not _ensure_mt5():
        return {"available": False}
    acct = mt5.account_info()
    if acct is None:
        _mt5_reset()
        return {"available": False}
    positions = mt5.positions_get() or []
    return {
        "available": True,
        "margin": round(acct.margin, 2),
        "margin_free": round(acct.margin_free, 2),
        "margin_level": round(acct.margin_level, 2) if acct.margin_level else acct.margin_level,
        "open_lots": round(sum(p.volume for p in positions), 4) if positions else 0.0,
    }


def _compliance(days: int, source: str | None = None) -> dict:
    """Stack-level (tradingview) risk-cap compliance, keyed off the orders
    table's `source` column. `source`, when given, restricts both rows to
    that source; "other" naturally yields empty (the orders table never
    records "other"-EA activity -- only this stack's own tradingview orders
    go through it)."""
    cutoff = _cutoff_iso(days)
    if source:
        risk_orders = ts.query(
            "SELECT trace_id, ts, symbol, requested_risk, status FROM orders "
            "WHERE requested_risk IS NOT NULL AND ts >= ? AND source = ? ORDER BY ts DESC",
            (cutoff, source),
        )
        rejections = ts.query(
            "SELECT trace_id, ts, symbol, error FROM orders "
            "WHERE status='rejected' AND error LIKE 'risk sizing%' AND ts >= ? AND source = ? ORDER BY ts DESC",
            (cutoff, source),
        )
    else:
        risk_orders = ts.query(
            "SELECT trace_id, ts, symbol, requested_risk, status FROM orders "
            "WHERE requested_risk IS NOT NULL AND ts >= ? ORDER BY ts DESC",
            (cutoff,),
        )
        rejections = ts.query(
            "SELECT trace_id, ts, symbol, error FROM orders "
            "WHERE status='rejected' AND error LIKE 'risk sizing%' AND ts >= ? ORDER BY ts DESC",
            (cutoff,),
        )
    max_risk = max((r["requested_risk"] for r in risk_orders), default=None)
    return {
        "cap_usd": EA_SHIM_RISK_USD,
        "risk_sized_orders": {"count": len(risk_orders), "rows": risk_orders[:50]},
        "risk_cap_rejections": {"count": len(rejections), "rows": rejections[:50]},
        "max_single_order_risk": max_risk,
        "max_vs_cap_pct": round(max_risk / EA_SHIM_RISK_USD * 100, 1) if (max_risk and EA_SHIM_RISK_USD) else None,
    }


def risk_panel(days: int = 30, source: str | None = None) -> dict:
    """equity_curve and live margin are account-level and always unfiltered
    (labeled "account-wide" client-side); only compliance is source-filtered."""
    return {
        "days": days,
        "equity_curve": _equity_curve(days),
        "live": _live_risk(),
        "compliance": _compliance(days, source),
    }


# --- monthly P/L calendar -----------------------------------------------


def calendar_month(month: str, stack_only: bool = True, source: str | None = None) -> dict:
    """`source`, when given, takes priority over `stack_only` and restricts
    the month to that single source; otherwise `stack_only` behaves as
    before (excludes "other" EA/manual activity when true)."""
    now = datetime.now(timezone.utc)
    try:
        year_s, mon_s = month.split("-", 1)
        year, mon = int(year_s), int(mon_s)
        if not (1 <= mon <= 12):
            raise ValueError
    except (ValueError, AttributeError):
        year, mon = now.year, now.month
    start = datetime(year, mon, 1, tzinfo=timezone.utc)
    end = datetime(year + 1, 1, 1, tzinfo=timezone.utc) if mon == 12 else datetime(year, mon + 1, 1, tzinfo=timezone.utc)

    rows = ts.query(
        "SELECT close_ts, profit, source FROM closed_trades WHERE close_ts >= ? AND close_ts < ?",
        (start.isoformat(), end.isoformat()),
    )
    if source:
        rows = [r for r in rows if r.get("source") == source]
    elif stack_only:
        rows = [r for r in rows if r.get("source") != "other"]

    daily: dict[str, dict] = {}
    for r in rows:
        d = (r.get("close_ts") or "")[:10]
        if not d:
            continue
        b = daily.setdefault(d, {"net": 0.0, "count": 0})
        b["net"] += r.get("profit") or 0.0
        b["count"] += 1

    days = []
    cur = start
    while cur < end:
        key = cur.date().isoformat()
        b = daily.get(key, {"net": 0.0, "count": 0})
        days.append({"date": key, "weekday": cur.weekday(), "net": round(b["net"], 2), "count": b["count"]})
        cur += timedelta(days=1)

    prev_ref = start - timedelta(days=1)
    return {
        "month": f"{year:04d}-{mon:02d}",
        "days": days,
        "total": round(sum(d["net"] for d in days), 2),
        "prev": f"{prev_ref.year:04d}-{prev_ref.month:02d}",
        "next": f"{end.year:04d}-{end.month:02d}",
        "stack_only": stack_only,
        "source": source or "",
    }


# --- digest text -----------------------------------------------------------


def _digest_period(days: int, source: str | None = None) -> dict:
    """`source`, when given, restricts signals/orders/closed-trade counts to
    that source (the `signals` table has no source column and is no longer
    populated by anything in this stack, so a given source always zeroes it).
    Equity/margin readouts stay account-level/unfiltered. Callers that omit
    `source` (the --digest-now CLI) see no behavior change."""
    cutoff = _cutoff_iso(days)
    if source:
        received = 0
    else:
        received = len(ts.query("SELECT 1 FROM signals WHERE ts >= ?", (cutoff,)))
    if source:
        order_rows = ts.query("SELECT status FROM orders WHERE ts >= ? AND source = ?", (cutoff, source))
        closed_rows = ts.query(
            "SELECT profit, source FROM closed_trades WHERE close_ts >= ? AND source = ?", (cutoff, source)
        )
    else:
        order_rows = ts.query("SELECT status FROM orders WHERE ts >= ?", (cutoff,))
        closed_rows = ts.query("SELECT profit, source FROM closed_trades WHERE close_ts >= ?", (cutoff,))
    executed = sum(1 for o in order_rows if o.get("status") in ("filled", "placed"))
    wins = sum(1 for c in closed_rows if (c.get("profit") or 0) >= 0)
    losses = len(closed_rows) - wins
    net = round(sum(c.get("profit") or 0 for c in closed_rows), 2)
    by_source: dict[str, float] = {}
    for c in closed_rows:
        s = c.get("source") or "other"
        by_source[s] = round(by_source.get(s, 0.0) + (c.get("profit") or 0.0), 2)
    if source:
        rej = ts.query(
            "SELECT COUNT(*) AS n FROM orders WHERE status='rejected' AND error LIKE 'risk sizing%' "
            "AND ts >= ? AND source = ?",
            (cutoff, source),
        )
    else:
        rej = ts.query(
            "SELECT COUNT(*) AS n FROM orders WHERE status='rejected' AND error LIKE 'risk sizing%' AND ts >= ?",
            (cutoff,),
        )
    risk_cap_events = rej[0]["n"] if rej else 0

    equity_now = None
    floating = None
    margin_level = None
    if _ensure_mt5():
        acct = mt5.account_info()
        if acct is not None:
            equity_now = acct.equity
            margin_level = round(acct.margin_level, 2) if acct.margin_level else acct.margin_level
            positions = mt5.positions_get() or []
            floating = round(sum(p.profit for p in positions), 2) if positions else 0.0
        else:
            _mt5_reset()

    snaps = ts.query(
        "SELECT equity FROM equity_snapshots WHERE ts >= ? ORDER BY ts ASC LIMIT 1", (cutoff,)
    )
    if snaps:
        equity_start = snaps[0]["equity"]
    elif equity_now is not None:
        equity_start = round(equity_now - net, 2)
    else:
        equity_start = None

    return {
        "days": days, "received": received, "executed": executed,
        "wins": wins, "losses": losses, "net": net, "by_source": by_source,
        "floating": floating, "equity_now": equity_now, "equity_start": equity_start,
        "margin_level": margin_level, "risk_cap_events": risk_cap_events,
    }


def _fmt_signed(v: float) -> str:
    return f"{'+' if v >= 0 else ''}{v:.2f}"


def build_digest_text(days: int, weekly: bool) -> str:
    """Plain-text digest (no markdown), printed to the console by the
    --digest-now CLI path. Pure function of the store's current state."""
    d = _digest_period(days)
    period_label = "Weekly (7-day)" if weekly else f"{days}-day"
    lines = [
        f"Rey Capital trade digest — {period_label} — {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
        f"Signals received: {d['received']} | executed: {d['executed']}",
        f"Trades closed: {d['wins']}W / {d['losses']}L | Net P/L: {_fmt_signed(d['net'])} USD",
    ]
    if d["by_source"]:
        parts = ", ".join(f"{k}: {_fmt_signed(v)}" for k, v in sorted(d["by_source"].items()))
        lines.append(f"By source: {parts}")
    if d["floating"] is not None:
        lines.append(f"Floating P/L now: {_fmt_signed(d['floating'])} USD")
    if d["equity_now"] is not None:
        eq_start_s = f"{d['equity_start']:.2f}" if d["equity_start"] is not None else "n/a"
        lines.append(f"Equity now: {d['equity_now']:.2f} USD (period start: {eq_start_s})")
    else:
        lines.append("Equity now: MT5 unavailable")
    if d["margin_level"] is not None:
        lines.append(f"Margin level: {d['margin_level']:.1f}%")
    lines.append(f"Risk-cap rejections this period: {d['risk_cap_events']}")
    return "\n".join(lines)


# --- meta persistence ---------------------------------------------------


def _meta_get(key: str) -> str | None:
    conn = ts.get_conn()
    if conn is None:
        return None
    try:
        row = conn.execute("SELECT value FROM meta WHERE key=?", (key,)).fetchone()
        return row["value"] if row else None
    except Exception:  # noqa: BLE001
        return None
    finally:
        try:
            conn.close()
        except Exception:
            pass


# ---------------------------------------------------------------------------
# Pipeline status -- "is anything broken" panel (/api/pipeline). Pulls
# together the heartbeat ea_shim writes (meta table, so a hung-but-alive
# process can't masquerade as healthy the way pid-liveness does) with live
# HTTP health checks of the Go services. Every read here is either the
# existing no-throw ts.query()/​_meta_get, or a short-timeout HTTP probe that
# never raises.
# ---------------------------------------------------------------------------


def _hb_age_sec(key: str) -> float | None:
    """Seconds since the given meta heartbeat key was last written, or None
    if it's missing/unparseable (treated as "down" by callers)."""
    raw = _meta_get(key)
    if not raw:
        return None
    try:
        stamp = datetime.fromisoformat(raw)
    except ValueError:
        return None
    if stamp.tzinfo is None:
        stamp = stamp.replace(tzinfo=timezone.utc)
    return max(0.0, (datetime.now(timezone.utc) - stamp).total_seconds())


def _http_probe(url: str, timeout: float = 1.5) -> tuple[int | None, dict | None]:
    """GET url. Returns (status, parsed-json-body-or-None); status is None
    when the endpoint is unreachable (refused/timed out/DNS failure) --
    never raises."""
    try:
        with urllib.request.urlopen(url, timeout=timeout) as resp:
            status = resp.status
            raw = resp.read()
    except urllib.error.HTTPError as exc:
        status = exc.code
        raw = exc.read()
    except Exception:
        return None, None
    try:
        body = json.loads(raw.decode())
    except Exception:
        body = None
    return status, body


def pipeline_status() -> dict:
    """One component per link in the signal -> MT5-order chain, each
    {name, state: ok|warn|down, detail}, plus an overall verdict. See the
    build-out plan's "Overall verdict" rule: RED whenever a signal posted
    right now would NOT reach MT5 (any blocking component down, dry-run on,
    or zero channels enabled); AMBER for non-blocking degradation; GREEN
    only when it would actually reach MT5."""
    components: list[dict] = []

    status, _ = _http_probe("http://127.0.0.1:8081/health")
    components.append({
        "name": "ingress", "state": "ok" if status == 200 else "down",
        "detail": "healthy" if status == 200 else "unreachable (127.0.0.1:8081)",
    })

    status, _ = _http_probe("http://127.0.0.1:8082/health")
    components.append({
        "name": "bridge", "state": "ok" if status == 200 else "down",
        "detail": "healthy" if status == 200 else "unreachable (127.0.0.1:8082)",
    })

    ea_age = _hb_age_sec("hb_ea_shim")
    ea_state_raw = _meta_get("hb_ea_shim_state")
    try:
        ea_state = json.loads(ea_state_raw) if ea_state_raw else {}
    except json.JSONDecodeError:
        ea_state = {}
    ea_fresh = ea_age is not None and ea_age < _HB_STALE_AFTER_SEC
    ea_mt5_ok = bool(ea_state.get("mt5"))
    if ea_fresh and ea_mt5_ok:
        components.append({"name": "ea-shim", "state": "ok", "detail": f"heartbeat {int(ea_age)}s ago"})
    else:
        components.append({
            "name": "ea-shim", "state": "down",
            "detail": "orders would be accepted but never executed",
        })

    if ea_fresh and ea_mt5_ok:
        account = ea_state.get("account")
        demo = ea_state.get("demo")
        components.append({
            "name": "mt5", "state": "ok",
            "detail": f"account {account}" + (" (DEMO)" if demo else " (LIVE)"),
            "account": account, "demo": bool(demo) if demo is not None else None,
        })
    else:
        components.append({
            "name": "mt5", "state": "down",
            "detail": "unknown — ea-shim heartbeat stale or missing",
            "account": None, "demo": None,
        })

    # The dry-run component doubles as the model for the banner's dry-run
    # switch, so it carries the raw flag and its provenance, not just prose.
    dry_meta = ts.get_dry_run_meta(_DRY_RUN_ENV_DEFAULT)
    dry_on = bool(dry_meta.get("dry_run"))
    components.append({
        "name": "dry-run",
        "state": "warn" if dry_on else "ok",
        "detail": ("signals are parsed but NOT sent to the broker" if dry_on
                   else "live — signals are sent to the broker"),
        "dry_run": dry_on,
        "source": dry_meta.get("source"),
        "updated_ts": dry_meta.get("updated_ts"),
    })

    by_name = {c["name"]: c for c in components}
    blocking = ("ingress", "bridge", "ea-shim", "mt5")
    reasons = [f"{name} down" for name in blocking if by_name[name]["state"] == "down"]
    if dry_on:
        reasons.append("dry-run is on")

    if reasons:
        verdict, headline = "red", "SIGNALS NOT REACHING MT5"
    else:
        verdict, headline = "green", "SIGNALS REACHING MT5"

    return {
        "verdict": verdict,
        "headline": headline,
        "reasons": reasons,
        "components": components,
        "updated": datetime.now(timezone.utc).isoformat(),
    }


# ---------------------------------------------------------------------------
# Message ledger -- the two operator pages that answer "what did the app pass
# over?" and "what did it actually do?".
#
# Signals arrive as TradingView webhooks and only exist here once ingress has
# accepted them and issued a trace_id -- so `orders` is their ledger (the
# `signals` table is legacy/unpopulated -- see _trace_id_channels below).
# ---------------------------------------------------------------------------

_ACTIONED_STATUSES = ("accepted", "placed", "filled")
_ACTION_LABEL = {
    "newsltplong": "TP/SL amended on open longs",
    "newsltpshort": "TP/SL amended on open shorts",
    "closelong": "closed longs",
    "closeshort": "closed shorts",
    "closelongopenshort": "reversed long → short",
    "closeshortopenlong": "reversed short → long",
}


def _ledger_window(days: int) -> str:
    return (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()


def _trace_id_channels(days: int) -> dict[str, str]:
    """Map trace_id -> originating channel name, for orders whose signal was
    recorded in the `signals` table. That table is no longer populated by
    anything in this stack, so this currently always returns {} -- the query
    is left in place (harmless against an empty table) rather than
    special-cased away, in case a future signal source populates it again.
    """
    out: dict[str, str] = {}
    for s in ts.query(
        "SELECT channel, trace_ids FROM signals "
        "WHERE ts >= ? AND trace_ids IS NOT NULL AND trace_ids != '[]'",
        (_ledger_window(days + 1),),
    ):
        try:
            ids = json.loads(s.get("trace_ids") or "[]")
        except (TypeError, ValueError):
            continue
        name = s.get("channel") or "direct"
        for tid in ids if isinstance(ids, list) else []:
            if tid:
                out[str(tid)] = name
    return out


def ignored_messages(days: int = 7, source: str | None = None, limit: int = 500) -> dict:
    """Messages that arrived and produced NO order at the broker."""
    cutoff = _ledger_window(days)
    rows: list[dict] = []

    if source in (None, "", "tradingview", "other"):
        where = ["ts >= ?", "(status = 'rejected' OR (error IS NOT NULL AND error != ''))"]
        params: list = [cutoff]
        if source:
            where.append("source = ?")
            params.append(source)
        params.append(limit)
        for r in ts.query(
            "SELECT ts, source, command, symbol, status, error, comment FROM orders "
            f"WHERE {' AND '.join(where)} ORDER BY ts DESC LIMIT ?",
            tuple(params),
        ):
            rows.append({
                "ts": r.get("ts"),
                "source": r.get("source") or "other",
                "origin": r.get("comment") or "webhook",
                "reason": r.get("error") or "rejected at the broker",
                "symbol": r.get("symbol") or "",
                "text": f"{r.get('command') or ''} {r.get('symbol') or ''}".strip(),
            })

    rows.sort(key=lambda r: r.get("ts") or "", reverse=True)
    return {
        "rows": rows[:limit],
        "days": days,
        "source": source or "",
        "truncated": len(rows) > limit,
        # TradingView alerts malformed enough for ingress to refuse outright
        # never receive a trace_id, so they are not in this store at all --
        # say so rather than implying the list is exhaustive.
        "note": (
            "TradingView alerts rejected by ingress before a trace_id was issued "
            "are not recorded here — see the ingress log for those."
            if source in (None, "", "tradingview", "other") else ""
        ),
    }


def actioned_messages(days: int = 7, source: str | None = None, limit: int = 500) -> dict:
    """Everything this app actually did at MT5 -- one row per order/amendment
    that reached the broker path, newest first."""
    cutoff = _ledger_window(days)
    where = ["ts >= ?", "status IN (%s)" % ",".join("?" * len(_ACTIONED_STATUSES))]
    params: list = [cutoff, *_ACTIONED_STATUSES]
    if source:
        where.append("source = ?")
        params.append(source)
    params.append(limit)

    trace_channel = _trace_id_channels(days)
    rows = []
    for r in ts.query(
        "SELECT ts, trace_id, source, command, symbol, volume, entry, sl, tp, "
        "status, broker_order_id, comment FROM orders "
        f"WHERE {' AND '.join(where)} ORDER BY ts DESC LIMIT ?",
        tuple(params),
    ):
        cmd = (r.get("command") or "").lower()
        rows.append({
            "ts": r.get("ts"),
            "source": r.get("source") or "other",
            "origin": r.get("comment") or "",
            "channel": trace_channel.get(r.get("trace_id") or "", ""),
            "action": _ACTION_LABEL.get(cmd, f"{cmd} order placed" if cmd else "order"),
            "command": cmd,
            "symbol": r.get("symbol") or "",
            "volume": r.get("volume"),
            "entry": r.get("entry"),
            "sl": r.get("sl"),
            "tp": r.get("tp"),
            "status": r.get("status") or "",
            "broker_order_id": r.get("broker_order_id") or "",
            "trace_id": r.get("trace_id") or "",
        })
    return {"rows": rows, "days": days, "source": source or ""}


def _post_to_ingress(body: str) -> tuple[int, str, str]:
    """POST one already-built webhook command body to ingress. The ONLY
    function in this file that talks to ingress -- INGRESS_PORT is
    independently configurable so tests can redirect it to a local stub
    without ever touching the real service. Returns (http_status,
    response_text, trace_id); trace_id is "" if the response wasn't the
    expected {"trace_id": ...} JSON shape."""
    url = _INGRESS_WEBHOOK_URL
    if INGRESS_PERIMETER_TOKEN:
        url += f"?token={INGRESS_PERIMETER_TOKEN}"
    req = urllib.request.Request(
        url, data=body.encode(), headers={"Content-Type": "text/plain"}, method="POST"
    )
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            status = resp.status
            raw = resp.read().decode(errors="replace")[:500]
    except urllib.error.HTTPError as exc:
        status = exc.code
        raw = exc.read().decode(errors="replace")[:500]
    trace_id = ""
    try:
        trace_id = json.loads(raw).get("trace_id", "") or ""
    except (json.JSONDecodeError, AttributeError):
        pass
    return status, raw, trace_id


# --- "Send a sample TradingView signal" -----------------------------------
#
# Builds a raw ExecRelay webhook body in exactly the wire format
# pine/EMA915_Pullback_Webhook.pine sends (license,COMMAND,symbol,key=value...)
# and POSTs it straight to ingress via _post_to_ingress -- so "Preview" /
# "Send" here exercises the real TradingView -> ingress -> bridge -> ea_shim
# path, not a simulation of it.

_WEBHOOK_TEST_COMMANDS = {"BUY", "SELL", "CLOSELONG", "CLOSESHORT"}
_WEBHOOK_TEST_OPENS = {"BUY", "SELL"}


def build_test_signal_command(fields: dict) -> tuple[str | None, str | None]:
    """Returns (command, error); exactly one is non-None."""
    command = str(fields.get("command") or "").strip().upper()
    if command not in _WEBHOOK_TEST_COMMANDS:
        return None, f"command must be one of {', '.join(sorted(_WEBHOOK_TEST_COMMANDS))}"

    symbol = str(fields.get("symbol") or "").strip()
    if not symbol:
        return None, "symbol is required"

    parts = [LICENSE_ID, command, symbol]

    if command in _WEBHOOK_TEST_OPENS:
        vol_lots = str(fields.get("vol_lots") or "").strip()
        if not vol_lots:
            return None, "vol_lots is required for BUY/SELL"
        parts.append(f"vol_lots={vol_lots}")

        risk_mode = str(fields.get("risk_mode") or "bypass").strip()
        if risk_mode == "bypass":
            parts.append("risk=0")  # exact lots, no $-risk resizing (matches the pine script)
        elif risk_mode == "custom":
            risk_val = str(fields.get("risk") or "").strip()
            if not risk_val:
                return None, "risk is required when risk_mode is 'custom'"
            parts.append(f"risk={risk_val}")
        elif risk_mode != "stack":
            return None, "risk_mode must be one of bypass, custom, stack"
        # risk_mode == "stack": omit risk entirely, let EA_SHIM_RISK_USD size it.

        sl = str(fields.get("sl") or "").strip()
        tp = str(fields.get("tp") or "").strip()
        if not sl or not tp:
            return None, "sl and tp are required for BUY/SELL"
        parts.append(f"sl={sl}")
        parts.append(f"tp={tp}")

    comment = str(fields.get("comment") or "webhook-test").strip()
    if comment:
        parts.append(f"comment={comment}")
    if WEBHOOK_SECRET:
        parts.append(f"secret={WEBHOOK_SECRET}")

    return ",".join(parts), None


def send_test_signal(payload: dict) -> tuple[int, dict]:
    """Operator-triggered TradingView webhook test. Safety gates: explicit
    confirm:true required to actually send, dry-run always refuses. Returns
    a preview-shaped body ({ok, commands, warnings, errors}) that the
    dashboard's resubmit-confirm modal renders."""
    warnings: list[str] = []
    if effective_dry_run():
        warnings.append("system is in DRY-RUN")
    if not LICENSE_ID:
        warnings.append("DASHBOARD_LICENSE_ID is not set — ingress will reject with license_rejected")

    command, err = build_test_signal_command(payload)
    if err:
        return 400, {"ok": False, "commands": [], "warnings": warnings, "errors": [err]}

    redacted = _redact(command)
    preview = {"ok": True, "commands": [redacted], "warnings": warnings, "errors": []}

    if payload.get("confirm") is not True:
        return 200, preview

    if effective_dry_run():
        return 409, {
            "ok": False,
            "error": "system is in DRY-RUN — test signal refused (turn dry-run off in the pipeline bar to place orders)",
            "preview": preview,
        }

    try:
        status, resp_text, trace_id = _post_to_ingress(command)
    except Exception as exc:  # noqa: BLE001 - a network hiccup must not crash the handler
        status, resp_text, trace_id = 0, str(exc), ""

    ok = status == 200
    log_txn(
        WEBHOOK_TEST_TXN_LOG,
        command=redacted,
        http_status=status,
        trace_id=trace_id,
        ok=ok,
    )
    return 200, {
        "ok": ok,
        "results": [{"http_status": status, "trace_id": trace_id, "response": resp_text[:300]}],
        "warnings": warnings,
    }


# --- weekly XLSX export --------------------------------------------------


def build_weekly_xlsx(days: int = 7, source: str | None = None) -> bytes | None:
    """Returns None if openpyxl isn't installed -- caller responds 501.
    `source`, when given, filters the Summary figures and the Closed Trades
    sheet to that source; the Scorecard sheet stays unfiltered (it is itself
    the source comparison) and Equity stays account-level."""
    try:
        from openpyxl import Workbook
    except ImportError:
        return None

    d = _digest_period(days, source)
    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    ws.append(["Rey Capital — Weekly Management Report"])
    ws.append(["Period (days)", days])
    ws.append(["Source filter", source or "All sources"])
    ws.append(["Generated (UTC)", datetime.now(timezone.utc).isoformat()])
    ws.append([])
    ws.append(["Signals received", d["received"]])
    ws.append(["Orders executed", d["executed"]])
    ws.append(["Trades closed — wins", d["wins"]])
    ws.append(["Trades closed — losses", d["losses"]])
    ws.append(["Net P/L (period)", d["net"]])
    for src, val in sorted(d["by_source"].items()):
        ws.append([f"Net P/L — {src}", val])
    ws.append(["Floating P/L now", d["floating"]])
    ws.append(["Equity now", d["equity_now"]])
    ws.append(["Equity at period start", d["equity_start"]])
    ws.append(["Margin level (%)", d["margin_level"]])
    ws.append(["Risk-cap rejections", d["risk_cap_events"]])

    cutoff = _cutoff_iso(days)
    if source:
        closed_rows = ts.query(
            "SELECT position_id, close_ts, symbol, side, volume, entry_price, close_price, "
            "profit, magic, comment, source FROM closed_trades WHERE close_ts >= ? AND source = ? "
            "ORDER BY close_ts DESC",
            (cutoff, source),
        )
    else:
        closed_rows = ts.query(
            "SELECT position_id, close_ts, symbol, side, volume, entry_price, close_price, "
            "profit, magic, comment, source FROM closed_trades WHERE close_ts >= ? ORDER BY close_ts DESC",
            (cutoff,),
        )
    ws2 = wb.create_sheet("Closed Trades")
    ct_header = [
        "position_id", "close_ts", "symbol", "side", "volume", "entry_price",
        "close_price", "profit", "magic", "comment", "source",
    ]
    ws2.append(ct_header)
    for r in closed_rows:
        ws2.append([r.get(h) for h in ct_header])

    sc = scorecard(days)
    ws3 = wb.create_sheet("Scorecard")
    sc_header = [
        "channel", "received", "posted", "rejected", "orders_executed",
        "orders_open_pending", "wins", "losses", "net_pl", "avg_r",
    ]
    ws3.append(sc_header)
    for r in sc["rows"]:
        ws3.append([r.get(h) for h in sc_header])

    snaps = ts.query(
        "SELECT ts, balance, equity, margin, margin_free, floating FROM equity_snapshots "
        "WHERE ts >= ? ORDER BY ts ASC",
        (cutoff,),
    )
    ws4 = wb.create_sheet("Equity")
    eq_header = ["ts", "balance", "equity", "margin", "margin_free", "floating"]
    ws4.append(eq_header)
    for r in snaps:
        ws4.append([r.get(h) for h in eq_header])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# UI — Rey Capital design system (mirrored from AccountManagementSystem)
# ---------------------------------------------------------------------------

LOGO_SVG = """<svg viewBox="0 0 89 89" fill="currentColor" xmlns="http://www.w3.org/2000/svg" aria-hidden="true">
<g transform="matrix(28.48369473913729,0,0,28.48369473913729,-26.717708028593407,-27.14496304221794)">
<polygon points="2.499,1.705 4.062,2.605 4.062,1.854 2.498,0.953 0.938,1.855 0.938,2.607"/>
<polygon points="3.812,3.51 4.062,3.363 4.062,2.902 2.498,2 0.938,2.902 0.938,3.365 1.188,3.51 2.498,2.752"/>
<polygon points="2.499,3.818 2.896,4.047 3.548,3.672 2.498,3.066 1.452,3.672 2.104,4.047"/>
</g></svg>"""

PAGE = """<!doctype html>
<html><head><meta charset="utf-8">
<script>(function(){try{var k="execrelay-dashboard-theme";var t=localStorage.getItem(k)||((window.matchMedia&&matchMedia("(prefers-color-scheme: light)").matches)?"light":"dark");if(t==="light")document.documentElement.classList.add("light");}catch(e){}})();</script>
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Rey Capital | Trade Operations</title>
<link rel="icon" href="/assets/favicon.png">
<style>
:root {
  --color-background: #020c1b;
  --color-surface: #061526;
  --color-surface-2: #0b2040;
  --color-border: #123060;
  --color-border-light: #1d4a8a;
  --color-primary: #00c2e0;
  --color-primary-dim: #00a3be;
  --color-primary-glow: rgb(0 194 224 / 0.18);
  --color-profit: #05e8a4;
  --color-profit-dim: rgb(5 232 164 / 0.14);
  --color-loss: #ff3d5f;
  --color-loss-dim: rgb(255 61 95 / 0.14);
  --color-warning: #ffb52e;
  --color-gold: #f4b942;
  --color-text: #cde4ff;
  --color-text-muted: #4e7aab;
  --color-text-dim: #7aa3cc;
  --radius-sm: 0.375rem;
  --radius: 0.625rem;
  --radius-lg: 0.875rem;
}
html.light {
  --color-background: #eef4ff;
  --color-surface: #ffffff;
  --color-surface-2: #deeaf8;
  --color-border: #b8d0ec;
  --color-border-light: #90b8e4;
  --color-primary: #0077b6;
  --color-primary-dim: #005f99;
  --color-primary-glow: rgb(0 119 182 / 0.12);
  --color-profit: #00916e;
  --color-profit-dim: rgb(0 145 110 / 0.12);
  --color-loss: #d62839;
  --color-loss-dim: rgb(214 40 57 / 0.12);
  --color-warning: #c77c00;
  --color-gold: #b8860b;
  --color-text: #08203f;
  --color-text-muted: #3a6290;
  --color-text-dim: #1e4e7a;
}
* { box-sizing: border-box; }
body {
  margin: 0;
  font-family: system-ui, "Segoe UI", Roboto, sans-serif;
  background-color: var(--color-background);
  color: var(--color-text);
}
.number { font-variant-numeric: tabular-nums; font-feature-settings: "tnum"; }
::-webkit-scrollbar { width: 4px; height: 4px; }
::-webkit-scrollbar-thumb { background: var(--color-border-light); border-radius: 2px; }

.topbar { position: sticky; top: 0; z-index: 20; }
header {
  display: grid; grid-template-columns: 1fr auto 1fr; align-items: center;
  gap: 1rem; padding: 0.875rem 1.5rem;
  background: color-mix(in srgb, var(--color-surface) 85%, transparent 15%);
  backdrop-filter: blur(12px); -webkit-backdrop-filter: blur(12px);
  box-shadow: 0 1px 0 color-mix(in srgb, var(--color-primary) 6%, transparent), 0 4px 16px rgb(0 0 0 / 0.25);
}
.controlsbar {
  display: flex; align-items: center; gap: 1.25rem; flex-wrap: wrap;
  padding: 0.5rem 1.5rem;
  background: color-mix(in srgb, var(--color-surface-2) 90%, transparent 10%);
  backdrop-filter: blur(12px); -webkit-backdrop-filter: blur(12px);
  border-bottom: 1px solid var(--color-border);
  box-shadow: 0 4px 16px rgb(0 0 0 / 0.2);
}
.ctrl-group { display: flex; align-items: center; gap: 0.5rem; }
.ctrl-label { font-size: 0.66rem; letter-spacing: 0.1em; text-transform: uppercase; color: var(--color-text-muted); }
.seg-select {
  background: var(--color-background); color: var(--color-text);
  border: 1px solid var(--color-border-light); border-radius: 999px;
  padding: 0.28rem 1.9rem 0.28rem 0.8rem; font-size: 0.74rem; cursor: pointer;
  font-family: inherit;
  appearance: none; -webkit-appearance: none; -moz-appearance: none;
  background-image: url("data:image/svg+xml;utf8,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 10 6'><path d='M1 1l4 4 4-4' fill='none' stroke='%234e7aab' stroke-width='1.4'/></svg>");
  background-repeat: no-repeat; background-position: right 0.7rem center; background-size: 0.6rem;
}
.seg-select:focus { outline: 1px solid var(--color-primary); }
.controlsbar .icon-btn { margin-left: auto; }
.brand { display: flex; align-items: center; gap: 0.65rem; }
.logo-tile {
  width: 2.25rem; height: 2.25rem; border-radius: 0.5rem; background: #004AAC;
  display: flex; align-items: center; justify-content: center; flex-shrink: 0;
  box-shadow: 0 4px 14px rgb(0 0 0 / 0.4);
}
.logo-tile svg { width: 58%; height: 58%; color: #fff; }
.brand-name { font-size: 0.95rem; font-weight: 600; line-height: 1.15; }
.brand-sub { font-size: 0.68rem; color: var(--color-text-muted); }
.kpi { text-align: center; }
.kpi-label { font-size: 10px; letter-spacing: 0.15em; text-transform: uppercase; color: var(--color-text-muted); }
.kpi-value { font-size: 1.125rem; font-weight: 600; }
.header-right { display: flex; align-items: center; gap: 0.75rem; justify-self: end; }
.icon-btn {
  background: transparent; border: 1px solid var(--color-border-light); color: var(--color-text-dim);
  border-radius: 999px; width: 2rem; height: 2rem; display: flex; align-items: center; justify-content: center;
  cursor: pointer; font-size: 0.95rem; line-height: 1; flex-shrink: 0;
}
.icon-btn:hover { background: var(--color-primary-glow); color: var(--color-primary); }
.acct { text-align: right; font-size: 0.72rem; color: var(--color-text-muted); line-height: 1.5; }
.acct b { color: var(--color-text-dim); font-weight: 500; }

main { max-width: 1600px; margin: 0 auto; padding: 1.5rem; }
.cards { display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 1rem; }
.srcpl { display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 1rem; margin-bottom: 1.5rem; }
.stat-card {
  background: linear-gradient(180deg, var(--color-surface-2), var(--color-surface));
  border: 1px solid var(--color-border); border-radius: var(--radius);
  padding: 1rem 1.25rem; transition: transform .15s, box-shadow .15s;
}
.stat-card:hover { transform: translateY(-1px); box-shadow: 0 6px 20px rgb(0 0 0 / 0.35), 0 0 12px var(--color-primary-glow); }
.stat-card b { display: block; font-size: 1.35rem; font-weight: 600; }
.stat-card span { color: var(--color-text-muted); font-size: 0.72rem; }
.stat-card .sub { font-size: 0.68rem; color: var(--color-text-dim); margin-top: 2px; }

h2 { font-size: 0.82rem; letter-spacing: 0.08em; text-transform: uppercase; color: var(--color-text-dim); margin: 2rem 0 0.75rem; display:flex; align-items:center; gap:.5rem; flex-wrap: wrap; }
h2 .chip { font-size: 0.65rem; letter-spacing: normal; text-transform: none; border-radius: 999px; padding: 0.1rem 0.6rem; border: 1px solid var(--color-border-light); color: var(--color-text-muted); }
.grid2 { display: grid; grid-template-columns: 1fr; gap: 1.5rem; }
@media (min-width: 1100px) { .grid2 { grid-template-columns: 1fr 1fr; } }
.grid3 { display: grid; grid-template-columns: 1fr; gap: 1.5rem; }
@media (min-width: 1100px) { .grid3 { grid-template-columns: 1fr 1fr 1.6fr; } }
.donut-card svg { width: 100%; height: 170px; display: block; overflow: visible; }
.donut-wrap { display: flex; align-items: center; justify-content: center; }
.donut-legend { display: flex; gap: 1rem; justify-content: center; margin-top: 0.5rem; font-size: 0.72rem; color: var(--color-text-dim); }
.donut-legend .dot { display: inline-block; width: 0.55rem; height: 0.55rem; border-radius: 999px; margin-right: 0.3rem; vertical-align: middle; }
.stat-card.highlight { border-color: var(--color-primary); box-shadow: 0 0 0 1px var(--color-primary) inset, 0 0 16px var(--color-primary-glow); }
.chip.acctwide { border-style: dashed; }

.seg { display: inline-flex; border: 1px solid var(--color-border-light); border-radius: 999px; overflow: hidden; }
.seg button { background: transparent; border: 0; color: var(--color-text-muted); font-size: 0.68rem; padding: 0.25rem 0.7rem; cursor: pointer; letter-spacing: normal; text-transform: none; }
.seg button.active { background: var(--color-primary-glow); color: var(--color-primary); }
.exports { font-size: 0.72rem; letter-spacing: normal; text-transform: none; }
.exports a { color: var(--color-primary); text-decoration: none; margin-right: 0.75rem; }
.exports a:hover { text-decoration: underline; }

.review-strip { display: flex; flex-wrap: wrap; gap: 0.5rem; margin: 0 0 1rem; }
.review-chip {
  display: flex; align-items: center; gap: 0.5rem; background: var(--color-surface-2);
  border: 1px solid var(--color-border); border-radius: 999px; padding: 0.3rem 0.4rem 0.3rem 0.75rem;
  font-size: 0.72rem; color: var(--color-text-dim);
}
.review-chip button {
  background: transparent; border: 1px solid var(--color-border-light); color: var(--color-primary);
  border-radius: 999px; padding: 0.1rem 0.6rem; font-size: 0.68rem; cursor: pointer;
}

.chart-card {
  background: linear-gradient(180deg, var(--color-surface-2), var(--color-surface));
  border: 1px solid var(--color-border); border-radius: var(--radius-lg);
  padding: 1rem 1.25rem; margin-bottom: 0.5rem; overflow-x: auto;
}
.chart-card svg { width: 100%; height: 120px; display: block; min-width: 320px; }
#chart-symbol svg { height: auto; min-height: 120px; }
#chart-winloss svg, #chart-buysell svg { height: 170px; width: 170px; min-width: 0; }

.tablewrap { overflow-x: auto; border: 1px solid var(--color-border); border-radius: var(--radius-lg); background: linear-gradient(180deg, color-mix(in srgb, var(--color-surface-2) 60%, var(--color-surface)), var(--color-surface)); }
table { width: 100%; border-collapse: collapse; font-size: 0.82rem; }
thead { background: var(--color-surface-2); }
th { text-align: left; padding: 0.5rem 0.75rem; color: var(--color-text-muted); font-weight: 500; font-size: 0.72rem; white-space: nowrap; }
td { padding: 0.45rem 0.75rem; border-top: 1px solid var(--color-border); white-space: nowrap; }
tfoot td { border-top: 1px solid var(--color-border-light); font-weight: 600; }
.pos { color: var(--color-profit); } .neg { color: var(--color-loss); } .warn { color: var(--color-warning); }
.muted { color: var(--color-text-muted); }
.badge { display: inline-block; padding: 0.05rem 0.5rem; border-radius: 999px; font-size: 0.68rem; }
.badge.buy { background: var(--color-profit-dim); color: var(--color-profit); }
.badge.sell { background: var(--color-loss-dim); color: var(--color-loss); }
.badge.tg { background: var(--color-primary-glow); color: var(--color-primary); }
.badge.tv { background: color-mix(in srgb, var(--color-gold) 15%, transparent); color: var(--color-gold); }
.badge.ok { background: var(--color-profit-dim); color: var(--color-profit); }
.badge.bad { background: var(--color-loss-dim); color: var(--color-loss); }
.badge.neutral { background: var(--color-surface-2); color: var(--color-text-dim); }
.stars { color: var(--color-gold); letter-spacing: 1px; }
button.jbtn {
  background: transparent; border: 1px solid var(--color-border-light); color: var(--color-primary);
  border-radius: var(--radius-sm); padding: 0.15rem 0.6rem; font-size: 0.7rem; cursor: pointer;
}
button.jbtn:hover { background: var(--color-primary-glow); }

#modal-scrim { position: fixed; inset: 0; background: rgb(0 0 0 / 0.55); display: none; align-items: center; justify-content: center; z-index: 50; }
#modal {
  width: min(480px, 92vw); background: linear-gradient(180deg, var(--color-surface-2), var(--color-surface));
  border: 1px solid var(--color-border-light); border-radius: var(--radius-lg); padding: 1.5rem;
  box-shadow: 0 20px 60px rgb(0 0 0 / 0.5);
}
#modal h3 { margin: 0 0 1rem; font-size: 0.95rem; }
#modal label { display: block; font-size: 0.7rem; color: var(--color-text-muted); margin: 0.7rem 0 0.25rem; text-transform: uppercase; letter-spacing: 0.06em; }
#modal input[type=text], #modal textarea, #modal select {
  width: 100%; background: var(--color-background); color: var(--color-text);
  border: 1px solid var(--color-border); border-radius: var(--radius-sm); padding: 0.45rem 0.6rem; font-size: 0.82rem;
  font-family: inherit;
}
#modal textarea { min-height: 70px; resize: vertical; }
.rating-row { display: flex; gap: 0.3rem; font-size: 1.3rem; cursor: pointer; color: var(--color-border-light); }
.rating-row span.on { color: var(--color-gold); }
.modal-actions { display: flex; justify-content: flex-end; gap: 0.6rem; margin-top: 1.2rem; }
.btn-primary {
  background: linear-gradient(135deg, var(--color-primary), var(--color-primary-dim)); color: #04121f;
  border: 0; border-radius: var(--radius-sm); padding: 0.45rem 1.1rem; font-weight: 600; font-size: 0.8rem; cursor: pointer;
}
.btn-ghost { background: transparent; border: 1px solid var(--color-border); color: var(--color-text-dim); border-radius: var(--radius-sm); padding: 0.45rem 1rem; font-size: 0.8rem; cursor: pointer; }
.checkline { display: flex; align-items: center; gap: 0.5rem; margin-top: 0.8rem; font-size: 0.8rem; color: var(--color-text-dim); }
#meta { color: var(--color-text-muted); font-size: 0.72rem; margin: 2.5rem 0 1rem; }

.compliance-strip { display: grid; grid-template-columns: repeat(auto-fit, minmax(170px, 1fr)); gap: 1rem; margin-bottom: 1rem; }
.compliance-strip .stat-card.bad-cap { border-color: var(--color-loss); box-shadow: 0 0 0 1px var(--color-loss) inset; }
.cal-controls { display: flex; align-items: center; gap: 0.6rem; }
.cal-controls button.navbtn {
  background: transparent; border: 1px solid var(--color-border-light); color: var(--color-primary);
  border-radius: var(--radius-sm); padding: 0.15rem 0.55rem; font-size: 0.75rem; cursor: pointer;
}
.cal-controls button.navbtn:hover { background: var(--color-primary-glow); }
.cal-controls .cal-label { font-size: 0.78rem; color: var(--color-text-dim); min-width: 6.5em; text-align: center; }
.cal-grid {
  display: grid; grid-template-columns: repeat(7, 1fr); gap: 0.4rem;
  background: linear-gradient(180deg, color-mix(in srgb, var(--color-surface-2) 60%, var(--color-surface)), var(--color-surface));
  border: 1px solid var(--color-border); border-radius: var(--radius-lg); padding: 0.85rem;
}
.cal-dow { font-size: 0.65rem; letter-spacing: 0.06em; text-transform: uppercase; color: var(--color-text-muted); text-align: center; padding-bottom: 0.15rem; }
.cal-cell {
  min-height: 58px; border-radius: var(--radius-sm); border: 1px solid var(--color-border);
  padding: 0.3rem 0.4rem; font-size: 0.72rem; background: var(--color-surface);
}
.cal-cell.empty { border: none; background: transparent; }
.cal-cell.win { background: var(--color-profit-dim); border-color: color-mix(in srgb, var(--color-profit) 35%, var(--color-border)); }
.cal-cell.loss { background: var(--color-loss-dim); border-color: color-mix(in srgb, var(--color-loss) 35%, var(--color-border)); }
.cal-cell .d { color: var(--color-text-muted); font-size: 0.68rem; }
.cal-cell .amt { font-weight: 600; margin-top: 0.15rem; }

.pipeline-banner { display: flex; align-items: center; gap: 1rem; flex-wrap: wrap; padding: 0.75rem 1.5rem; border-bottom: 1px solid var(--color-border); background: color-mix(in srgb, var(--color-surface) 92%, transparent); }
.pipeline-pill { display: inline-flex; align-items: center; gap: 0.5rem; font-weight: 700; font-size: 0.85rem; letter-spacing: 0.02em; padding: 0.5rem 1.15rem; border-radius: 999px; white-space: nowrap; }
.pipeline-pill.green { background: var(--color-profit-dim); color: var(--color-profit); border: 1px solid color-mix(in srgb, var(--color-profit) 40%, transparent); }
.pipeline-pill.amber { background: rgb(255 181 46 / 0.16); color: var(--color-warning); border: 1px solid color-mix(in srgb, var(--color-warning) 45%, transparent); }
.pipeline-pill.red { background: var(--color-loss-dim); color: var(--color-loss); border: 1px solid color-mix(in srgb, var(--color-loss) 55%, transparent); animation: pipeline-pulse-red 2s ease-in-out infinite; }
@keyframes pipeline-pulse-red { 0%, 100% { box-shadow: 0 0 8px color-mix(in srgb, var(--color-loss) 25%, transparent); } 50% { box-shadow: 0 0 22px color-mix(in srgb, var(--color-loss) 60%, transparent); } }
.pipeline-chips { display: flex; gap: 0.5rem; flex-wrap: wrap; }
.pipeline-chip { display: inline-flex; align-items: center; gap: 0.4rem; font-size: 0.72rem; padding: 0.28rem 0.7rem; border-radius: 999px; border: 1px solid var(--color-border-light); }
.pipeline-chip .dot { width: 0.5rem; height: 0.5rem; border-radius: 999px; background: currentColor; flex-shrink: 0; }
.pipeline-chip.ok { color: var(--color-profit); }
.pipeline-chip.warn { color: var(--color-warning); }
.pipeline-chip.down { color: var(--color-loss); }
.dryrun-control { display: inline-flex; align-items: center; gap: 0.55rem; margin-left: auto; padding: 0.3rem 0.8rem; border-radius: 999px; border: 1px solid var(--color-border-light); background: var(--color-surface-2); }
.dryrun-control[hidden] { display: none; }
.dryrun-control .dryrun-label { font-size: 0.72rem; font-weight: 600; text-transform: uppercase; letter-spacing: 0.06em; color: var(--color-text-muted); }
.dryrun-control .dryrun-state { font-size: 0.72rem; font-weight: 700; white-space: nowrap; }
.dryrun-control.on .dryrun-state { color: var(--color-warning); }
.dryrun-control.off .dryrun-state { color: var(--color-gold); }
.dryrun-control.busy { opacity: 0.55; pointer-events: none; }

.page-tabs { display: flex; gap: 0.25rem; padding: 0 1.5rem; border-bottom: 1px solid var(--color-border); background: color-mix(in srgb, var(--color-surface) 92%, transparent); }
.page-tabs button { appearance: none; background: none; border: none; border-bottom: 2px solid transparent; color: var(--color-text-muted); font-size: 0.8rem; font-weight: 600; letter-spacing: 0.02em; padding: 0.7rem 1rem; cursor: pointer; }
.page-tabs button:hover { color: var(--color-text); }
.page-tabs button.active { color: var(--color-accent, var(--color-text)); border-bottom-color: currentColor; }
/* Page switching is CSS-only so it never fights the inline display rules the
   overview's own refreshes set -- switching back reveals exactly what they left. */
main[data-page="overview"] > .ledger-page { display: none !important; }
main[data-page="ignored"]  > *:not(#page-ignored)  { display: none !important; }
main[data-page="actioned"] > *:not(#page-actioned) { display: none !important; }
.ledger-text { white-space: pre-wrap; word-break: break-word; max-width: 46ch; font-size: 0.76rem; color: var(--color-text-dim); }
.ledger-reason { white-space: normal; max-width: 24ch; }
.badge.live { background: color-mix(in srgb, var(--color-gold) 22%, transparent); color: var(--color-gold); font-weight: 700; }
.badge.demo { background: var(--color-surface-2); color: var(--color-text-dim); }

.toggle-switch { position: relative; display: inline-block; width: 2.2rem; height: 1.25rem; vertical-align: middle; }
.toggle-switch input { opacity: 0; width: 0; height: 0; }
.toggle-slider { position: absolute; inset: 0; background: var(--color-border-light); border-radius: 999px; cursor: pointer; transition: background .15s; }
.toggle-slider:before { content: ""; position: absolute; width: 0.95rem; height: 0.95rem; left: 0.15rem; top: 0.15rem; background: #fff; border-radius: 50%; transition: transform .15s; }
.toggle-switch input:checked + .toggle-slider { background: var(--color-profit); }
.toggle-switch input:checked + .toggle-slider:before { transform: translateX(0.95rem); }
.toggle-switch input:disabled + .toggle-slider { opacity: 0.5; cursor: default; }

.addchannel-form { display: flex; gap: 0.75rem; flex-wrap: wrap; align-items: flex-end; margin-bottom: 1rem; }
.addchannel-form > div { display: flex; flex-direction: column; }
.addchannel-form label { font-size: 0.66rem; color: var(--color-text-muted); text-transform: uppercase; letter-spacing: 0.06em; margin-bottom: 0.3rem; }
.addchannel-form input, .addchannel-form select { background: var(--color-background); color: var(--color-text); border: 1px solid var(--color-border); border-radius: var(--radius-sm); padding: 0.42rem 0.6rem; font-size: 0.8rem; font-family: inherit; min-width: 14rem; }
.addchannel-form .hint { font-size: 0.7rem; color: var(--color-warning); align-self: center; }
.unregistered-row { border-left: 3px solid var(--color-warning); }
.btn-danger { background: linear-gradient(135deg, var(--color-loss), #a4152b); color: #fff; border: 0; border-radius: var(--radius-sm); padding: 0.45rem 1.1rem; font-weight: 600; font-size: 0.8rem; cursor: pointer; }

.empty-state {
  padding: 1.35rem 1rem; text-align: center; color: var(--color-text-dim); font-size: 0.82rem;
  border: 1px dashed var(--color-border-light); border-radius: var(--radius-lg);
  background: linear-gradient(180deg, color-mix(in srgb, var(--color-surface-2) 40%, var(--color-surface)), var(--color-surface));
}
.manual-signal-box { margin-top: 1.25rem; }
.manual-signal-box h3 { font-size: 0.78rem; letter-spacing: 0.06em; text-transform: uppercase; color: var(--color-text-dim); margin: 0 0 0.4rem; }
.manual-signal-box textarea {
  width: 100%; min-height: 90px; background: var(--color-background); color: var(--color-text);
  border: 1px solid var(--color-border); border-radius: var(--radius-sm); padding: 0.55rem 0.7rem;
  font-size: 0.8rem; font-family: ui-monospace, Consolas, monospace; resize: vertical;
}
button.actbtn {
  background: transparent; border: 1px solid var(--color-border-light); color: var(--color-primary);
  border-radius: var(--radius-sm); padding: 0.18rem 0.6rem; font-size: 0.72rem; cursor: pointer; margin-right: 0.3rem;
}
button.actbtn:hover { background: var(--color-primary-glow); }
</style></head><body>
<div class="topbar">
<header>
  <div class="brand">
    <div class="logo-tile">__LOGO__</div>
    <div><div class="brand-name">Rey Capital</div><div class="brand-sub">Trade Operations</div></div>
  </div>
  <div class="kpi"><div class="kpi-label" id="kpi-label">Net P/L · 7 days</div><div class="kpi-value number" id="kpi-net">—</div></div>
  <div class="header-right">
    <div class="acct" id="acct">connecting…</div>
  </div>
</header>
<div class="controlsbar">
  <div class="ctrl-group">
    <span class="ctrl-label">Source</span>
    <select id="source-filter" class="seg-select">
      <option value="">All sources</option>
      <option value="tradingview">TradingView</option>
      <option value="other">Other EAs</option>
    </select>
  </div>
  <div class="ctrl-group">
    <span class="ctrl-label">Window</span>
    <span class="seg" id="seg-days">
      <button type="button" data-days="7" class="active">7d</button>
      <button type="button" data-days="30">30d</button>
      <button type="button" data-days="90">90d</button>
    </span>
  </div>
  <button id="theme-toggle" class="icon-btn" type="button" title="Toggle theme" aria-label="Toggle theme">&#9789;</button>
</div>
</div>
<div class="pipeline-banner">
  <div class="pipeline-pill" id="pipeline-pill">checking pipeline…</div>
  <div class="pipeline-chips" id="pipeline-chips"></div>
  <div class="dryrun-control" id="dryrun-control" hidden>
    <span class="dryrun-label">Dry run</span>
    <label class="toggle-switch"><input type="checkbox" id="dryrun-toggle" aria-label="Dry run mode"><span class="toggle-slider"></span></label>
    <span class="dryrun-state" id="dryrun-state"></span>
  </div>
</div>
<nav class="page-tabs" id="page-tabs">
  <button type="button" data-page="overview" class="active">Overview</button>
  <button type="button" data-page="ignored">Ignored messages</button>
  <button type="button" data-page="actioned">Actioned in MT5</button>
</nav>
<main data-page="overview">
  <div class="cards" id="cards"></div>

  <section id="section-webhook-test">
    <h2>Send a sample TradingView signal</h2>
    <p class="muted" style="font-size:0.76rem;margin:0 0 .5rem">
      Builds a webhook body in the exact format the TradingView strategy sends and posts it to
      the real ingress endpoint — the same path a live TradingView alert takes. This can place a
      real order; nothing is sent until you confirm, and it's refused outright while dry-run is on.
    </p>
    <form class="addchannel-form" id="webhook-test-form" onsubmit="return false">
      <div>
        <label>Command</label>
        <select id="wt-command">
          <option value="BUY">Buy</option>
          <option value="SELL">Sell</option>
          <option value="CLOSELONG">Close long</option>
          <option value="CLOSESHORT">Close short</option>
        </select>
      </div>
      <div>
        <label>Symbol</label>
        <input type="text" id="wt-symbol" value="XAUUSD" style="min-width:8rem">
      </div>
      <div id="wt-open-fields" style="display:contents">
        <div>
          <label>Volume (lots)</label>
          <input type="text" id="wt-vol-lots" value="0.01" style="min-width:6rem">
        </div>
        <div>
          <label>Risk sizing</label>
          <select id="wt-risk-mode">
            <option value="bypass">Exact lots (risk=0)</option>
            <option value="stack">Stack $-risk sizing</option>
            <option value="custom">Custom $ risk</option>
          </select>
        </div>
        <div id="wt-risk-custom-field" style="display:none">
          <label>Risk ($)</label>
          <input type="text" id="wt-risk" value="25" style="min-width:6rem">
        </div>
        <div>
          <label>Stop loss (price)</label>
          <input type="text" id="wt-sl" placeholder="e.g. 4398" style="min-width:8rem">
        </div>
        <div>
          <label>Take profit (price)</label>
          <input type="text" id="wt-tp" placeholder="e.g. 4368" style="min-width:8rem">
        </div>
      </div>
      <div>
        <label>Comment</label>
        <input type="text" id="wt-comment" value="webhook-test" style="min-width:8rem">
      </div>
      <button class="btn-primary" type="button" id="webhook-test-preview-btn">Preview</button>
    </form>
  </section>

  <h2>Performance</h2>
  <div class="grid2">
    <div class="chart-card"><h2 style="margin:0 0 .5rem" id="chart-equity-h2">Equity curve <span class="chip acctwide" id="chip-equity">account-wide</span></h2><div id="chart-equity"></div></div>
    <div class="chart-card"><h2 style="margin:0 0 .5rem">Cumulative net P/L <span class="chip" id="chip-cum"></span></h2><div id="chart-cum"></div></div>
  </div>
  <div class="grid2">
    <div class="chart-card"><h2 style="margin:0 0 .5rem">Daily P/L</h2><div id="chart-daily"></div></div>
    <div class="chart-card"><h2 style="margin:0 0 .5rem">P/L by symbol <span class="chip" id="chip-symbol"></span></h2><div id="chart-symbol"></div></div>
  </div>
  <div class="grid3">
    <div class="chart-card donut-card"><h2 style="margin:0 0 .5rem">Win / loss</h2><div class="donut-wrap" id="chart-winloss"></div></div>
    <div class="chart-card donut-card"><h2 style="margin:0 0 .5rem">Buy / sell split</h2><div class="donut-wrap" id="chart-buysell"></div></div>
    <div class="chart-card">
      <h2 style="margin:0 0 .5rem">Monthly P/L calendar <span class="chip" id="chip-calendar"></span>
        <span class="cal-controls">
          <button type="button" class="navbtn" id="cal-prev">&#8592;</button>
          <span class="cal-label" id="cal-label">—</span>
          <button type="button" class="navbtn" id="cal-next">&#8594;</button>
        </span>
        <span class="seg" id="seg-cal-source">
          <button type="button" data-source="stack" class="active">Stack only</button>
          <button type="button" data-source="all">All sources</button>
        </span>
      </h2>
      <div id="cal-grid" class="cal-grid"></div>
    </div>
  </div>

  <h2>Performance by source <span class="chip" id="chip-srcpl"></span></h2>
  <div class="srcpl" id="srcpl"></div>

  <h2>Channel scorecard <span class="chip" id="chip-scorecard"></span></h2>
  <div class="tablewrap"><table id="tbl-scorecard"></table></div>

  <h2>Risk &amp; exposure <span class="chip" id="chip-risk"></span></h2>
  <div class="cards" id="risk-cards" style="margin-bottom:1rem"></div>
  <div class="compliance-strip" id="compliance-strip"></div>
  <div class="tablewrap" id="wrap-risk-rejections"><table id="tbl-risk-rejections"></table></div>

  <div class="grid2">
    <section>
      <h2>Orders executed <span class="chip" id="chip-orders"></span></h2>
      <div class="tablewrap"><table id="tbl-orders"></table></div>
    </section>
    <section>
      <h2>Open positions <span class="chip" id="chip-open"></span></h2>
      <div class="tablewrap"><table id="tbl-open"></table></div>
    </section>
  </div>

  <h2>Closed trades &amp; journal <span class="chip" id="chip-journal"></span>
    <span class="exports">
      <a id="export-trades" href="/api/export/trades.csv">Export trades CSV</a><a id="export-journal" href="/api/export/journal.csv">Export journal CSV</a><a id="export-weekly-xlsx" href="/api/export/weekly.xlsx">Export weekly XLSX</a>
    </span>
  </h2>
  <div class="review-strip" id="review-strip" style="display:none"></div>
  <div class="tablewrap"><table id="tbl-closed"></table></div>

  <div id="meta"></div>
  <section class="ledger-page" id="page-ignored">
    <h2>Ignored messages <span class="chip" id="chip-ignored"></span></h2>
    <p class="muted" style="font-size:0.8rem;margin:0 0 .75rem">
      Everything that arrived from an <b>enabled</b> source and produced no order at the broker.
      A channel disabled in the registry is not listed — it is not being read at all.
    </p>
    <div id="ignored-note" class="review-strip"></div>
    <div class="tablewrap"><table id="tbl-ignored"></table></div>
    <p class="muted" id="ignored-empty" style="font-size:0.8rem"></p>
  </section>

  <section class="ledger-page" id="page-actioned">
    <h2>Actioned in MT5 <span class="chip" id="chip-actioned"></span></h2>
    <p class="muted" style="font-size:0.8rem;margin:0 0 .75rem">
      Every order and amendment this app actually sent to the broker, newest first.
    </p>
    <div class="tablewrap"><table id="tbl-actioned"></table></div>
    <p class="muted" id="actioned-empty" style="font-size:0.8rem"></p>
  </section>
</main>

<div id="modal-scrim"><div id="modal">
  <h3 id="modal-title">Journal</h3>
  <input type="hidden" id="j-ticket">
  <label>Setup / pattern</label><input type="text" id="j-setup" placeholder="e.g. breakout retest, supply zone">
  <label>Emotion</label><select id="j-emotion"><option value="">—</option></select>
  <label>Mistakes</label><input type="text" id="j-mistakes" placeholder="comma-separated, e.g. chased entry, moved SL">
  <label>Rating</label><div class="rating-row" id="j-rating"></div>
  <label>Notes / lessons</label><textarea id="j-notes"></textarea>
  <div class="checkline"><input type="checkbox" id="j-reviewed"><label for="j-reviewed" style="margin:0;text-transform:none;letter-spacing:0">Reviewed</label></div>
  <div class="modal-actions">
    <button class="btn-ghost" type="button" id="j-cancel">Cancel</button>
    <button class="btn-primary" type="button" id="j-save">Save entry</button>
  </div>
</div></div>

<div id="dryrun-scrim" style="position:fixed;inset:0;background:rgb(0 0 0 / 0.55);display:none;align-items:center;justify-content:center;z-index:70">
  <div style="width:min(460px,92vw);background:linear-gradient(180deg,var(--color-surface-2),var(--color-surface));border:1px solid var(--color-border-light);border-radius:var(--radius-lg);padding:1.5rem">
    <h3 style="margin:0 0 .75rem;font-size:0.95rem">Turn dry-run OFF?</h3>
    <p class="muted" style="font-size:0.82rem;margin:0 0 .5rem">The next "Send a sample TradingView signal" you send from this dashboard will place a <b>real order</b> on the connected account. The change applies immediately — no restart needed.</p>
    <p style="font-size:0.85rem;font-weight:600" id="dryrun-acct"></p>
    <div class="modal-actions">
      <button class="btn-ghost" type="button" id="dryrun-cancel">Cancel</button>
      <button class="btn-danger" type="button" id="dryrun-confirm">Turn dry-run off</button>
    </div>
  </div>
</div>

<div id="resubmit-scrim" style="position:fixed;inset:0;background:rgb(0 0 0 / 0.55);display:none;align-items:center;justify-content:center;z-index:60">
  <div style="width:min(640px,94vw);max-height:88vh;overflow:auto;background:linear-gradient(180deg,var(--color-surface-2),var(--color-surface));border:1px solid var(--color-border-light);border-radius:var(--radius-lg);padding:1.5rem">
    <h3 style="margin:0 0 .75rem;font-size:0.95rem" id="resubmit-title">Place order(s)?</h3>
    <div id="resubmit-warnings"></div>
    <label style="display:block;font-size:0.7rem;color:var(--color-text-muted);margin:0.7rem 0 0.25rem;text-transform:uppercase;letter-spacing:0.06em">Exact commands (secret redacted)</label>
    <pre id="resubmit-commands" style="margin:0;padding:0.75rem;white-space:pre-wrap;word-break:break-all;font-size:0.74rem;font-family:ui-monospace,Consolas,monospace;background:var(--color-background);border:1px solid var(--color-border);border-radius:var(--radius-sm);max-height:220px;overflow:auto"></pre>
    <div id="resubmit-results" style="margin-top:0.75rem;font-size:0.8rem"></div>
    <div class="modal-actions">
      <button class="btn-ghost" type="button" id="resubmit-cancel">Cancel</button>
      <button class="btn-danger" type="button" id="resubmit-confirm" disabled>Place these orders</button>
    </div>
  </div>
</div>

<script>
const $ = id => document.getElementById(id);
const money = (v, c) => (v >= 0 ? "+" : "\u2212") + "$" + Math.abs(v).toFixed(2);
const cls = v => v >= 0 ? "pos" : "neg";
const esc = s => String(s ?? "").replace(/[&<>"']/g, ch => ({"&":"&amp;","<":"&lt;",">":"&gt;",'"':"&quot;","'":"&#39;"}[ch]));
const srcBadge = s => s === "tradingview" ? '<span class="badge tv">TradingView</span>'
  : '<span class="badge neutral">Other EA</span>';
const sideBadge = s => s === "close" ? '<span class="badge neutral">CLOSE</span>'
  : `<span class="badge ${s}">${s.toUpperCase()}</span>`;
const sourceLabel = s => s === "tradingview" ? "TradingView" : s === "other" ? "Other EAs" : "All sources";
let lastSummary = null, ratingVal = 0, currentDays = 7, currentSource = "";

// --- global filter state: persisted to localStorage + the URL hash so a
// reload (or a shared link) keeps the same days/source selection. -------
const STATE_DAYS_KEY = "execrelay-dashboard-days";
const STATE_SOURCE_KEY = "execrelay-dashboard-source";
const VALID_DAYS = [7, 30, 90];
const VALID_SOURCES = ["", "tradingview", "other"];

function loadState() {
  let days = null, source = null;
  try {
    const hash = new URLSearchParams(location.hash.replace(/^#/, ""));
    if (hash.has("days")) days = parseInt(hash.get("days"), 10);
    if (hash.has("source")) source = hash.get("source");
  } catch (e) {}
  if (days === null || !VALID_DAYS.includes(days)) {
    try { const v = parseInt(localStorage.getItem(STATE_DAYS_KEY), 10); if (VALID_DAYS.includes(v)) days = v; } catch (e) {}
  }
  if (source === null || !VALID_SOURCES.includes(source)) {
    try { const v = localStorage.getItem(STATE_SOURCE_KEY); if (VALID_SOURCES.includes(v)) source = v; } catch (e) {}
  }
  currentDays = VALID_DAYS.includes(days) ? days : 7;
  currentSource = VALID_SOURCES.includes(source) ? source : "";
}
function persistState() {
  try {
    localStorage.setItem(STATE_DAYS_KEY, String(currentDays));
    localStorage.setItem(STATE_SOURCE_KEY, currentSource);
  } catch (e) {}
  const params = new URLSearchParams();
  params.set("days", String(currentDays));
  if (currentSource) params.set("source", currentSource);
  try { history.replaceState(null, "", "#" + params.toString()); } catch (e) {}
}
function sourceQS() { return currentSource ? "&source=" + encodeURIComponent(currentSource) : ""; }

const TOKEN = new URLSearchParams(location.search).get("token") || "";
function authHeaders(extra) {
  const h = Object.assign({}, extra || {});
  if (TOKEN) h["X-Dashboard-Token"] = TOKEN;
  return h;
}
function withToken(url) {
  if (!TOKEN) return url;
  return url + (url.includes("?") ? "&" : "?") + "token=" + encodeURIComponent(TOKEN);
}
if (TOKEN) {
  const icon = document.querySelector('link[rel="icon"]');
  if (icon) icon.href = withToken("/assets/favicon.png");
}

function card(label, value, extra, sub, wrapCls) {
  return `<div class="stat-card${wrapCls?" "+wrapCls:""}"><b class="number ${extra||""}">${value}</b><span>${label}</span>${sub?`<div class="sub">${sub}</div>`:""}</div>`;
}
function table(id, header, rows, footer) {
  $(id).innerHTML = "<thead><tr>" + header.map(h => `<th>${h}</th>`).join("") + "</tr></thead><tbody>" +
    (rows.length ? rows.join("") : `<tr><td colspan="${header.length}" class="muted">none yet</td></tr>`) + "</tbody>" +
    (footer ? "<tfoot>" + footer + "</tfoot>" : "");
}

async function refresh() {
  await Promise.all([refreshSummary(), refreshScorecard(), refreshCalendar(), refreshPipeline(), refreshLedger()]);
}

// --- ledger pages: "Ignored messages" / "Actioned in MT5" ----------------
// Both follow the SOURCE dropdown and the day window, exactly like the
// overview: pick TradingView and only TradingView rows remain.

let currentPage = "overview";

function showPage(name) {
  currentPage = name;
  document.querySelector("main").dataset.page = name;
  for (const b of document.querySelectorAll("#page-tabs button")) {
    b.classList.toggle("active", b.dataset.page === name);
  }
  try { localStorage.setItem("execrelay-dashboard-page", name); } catch (e) {}
  refreshLedger();
}

function ledgerTime(ts) { return esc((ts || "").slice(0, 19).replace("T", " ")) || "—"; }

function sourceBadge(s) {
  const label = s === "tradingview" ? "TradingView" : "Other";
  return `<span class="badge neutral">${esc(label)}</span>`;
}

async function refreshIgnored() {
  const res = await fetch(withToken("/api/ignored?days=" + currentDays + sourceQS()), { headers: authHeaders() });
  if (!res.ok) return;
  const d = await res.json();
  const rows = d.rows || [];
  $("chip-ignored").textContent = `${rows.length}${d.truncated ? "+" : ""} over ${currentDays}d`;
  $("ignored-note").innerHTML = d.note ? `<div class="muted" style="font-size:0.76rem">${esc(d.note)}</div>` : "";
  $("ignored-empty").textContent = rows.length ? "" : "Nothing was ignored in this window.";
  table("tbl-ignored", ["When", "Source", "From", "Why it was ignored", "Symbol", "Message"], rows.map(r => `<tr>
    <td class="muted number">${ledgerTime(r.ts)}</td>
    <td>${sourceBadge(r.source)}</td>
    <td>${esc(r.origin || "—")}</td>
    <td class="ledger-reason">${esc(r.reason || "")}</td>
    <td>${esc(r.symbol || "—")}</td>
    <td class="ledger-text">${esc(r.text || "")}</td>
  </tr>`));
}

async function refreshActioned() {
  const res = await fetch(withToken("/api/actioned?days=" + currentDays + sourceQS()), { headers: authHeaders() });
  if (!res.ok) return;
  const d = await res.json();
  const rows = d.rows || [];
  $("chip-actioned").textContent = `${rows.length} over ${currentDays}d`;
  $("actioned-empty").textContent = rows.length ? "" : "Nothing has been sent to the broker in this window.";
  const num = v => (v === null || v === undefined || v === "") ? "—" : esc(String(v));
  table("tbl-actioned", ["When", "Source", "From", "Action", "Symbol", "Lots", "Entry", "SL", "TP", "Status", "Broker ID", "Tag"], rows.map(r => `<tr>
    <td class="muted number">${ledgerTime(r.ts)}</td>
    <td>${sourceBadge(r.source)}</td>
    <td>${esc(r.channel || "—")}</td>
    <td>${esc(r.action || "")}</td>
    <td>${esc(r.symbol || "—")}</td>
    <td class="number">${num(r.volume)}</td>
    <td class="number">${num(r.entry)}</td>
    <td class="number">${num(r.sl)}</td>
    <td class="number">${num(r.tp)}</td>
    <td>${esc(r.status || "")}</td>
    <td class="muted number">${esc(r.broker_order_id || "—")}</td>
    <td class="muted">${esc(r.origin || "—")}</td>
  </tr>`));
}

async function refreshLedger() {
  // Only fetch the page actually on screen -- these are the widest queries in
  // the dashboard, and the overview does not show either of them.
  try {
    if (currentPage === "ignored") await refreshIgnored();
    else if (currentPage === "actioned") await refreshActioned();
  } catch (e) {}
}

document.getElementById("page-tabs").addEventListener("click", (e) => {
  const btn = e.target.closest("button[data-page]");
  if (btn) showPage(btn.dataset.page);
});

// --- pipeline status banner --------------------------------------------

function pipelineChip(c) {
  const stateCls = c.state === "ok" ? "ok" : (c.state === "warn" ? "warn" : "down");
  let extra = "";
  if (c.name === "mt5" && c.account) {
    extra = ` <span class="badge ${c.demo ? "demo" : "live"}">${c.demo ? "DEMO" : "LIVE"}</span> #${esc(c.account)}`;
  }
  return `<span class="pipeline-chip ${stateCls}" title="${esc(c.detail || "")}"><span class="dot"></span>${esc(c.name)}${extra}</span>`;
}

async function refreshPipeline() {
  try {
    const res = await fetch(withToken("/api/pipeline"), { headers: authHeaders() });
    if (!res.ok) return;
    const p = await res.json();
    const pill = $("pipeline-pill");
    pill.className = "pipeline-pill " + p.verdict;
    pill.textContent = p.headline + (p.reasons && p.reasons.length ? " — " + p.reasons.join("; ") : "");
    $("pipeline-chips").innerHTML = p.components.map(pipelineChip).join("");
    renderDryRun(p.components.find(c => c.name === "dry-run"));
    const mt5 = p.components.find(c => c.name === "mt5");
    $("dryrun-acct").textContent = (mt5 && mt5.account)
      ? "Account #" + mt5.account + (mt5.demo === false ? " — REAL MONEY" : (mt5.demo ? " (demo account)" : ""))
      : "";
  } catch (e) {}
}

// --- dry-run switch ------------------------------------------------------
// The actionable version of the banner's dry-run chip. Turning dry-run OFF
// arms live trading, so that direction is held back behind a confirm modal
// and sends confirm:true -- the server refuses the change without it.

let dryRunApplying = false;

function renderDryRun(c) {
  const wrap = $("dryrun-control");
  if (!c) { wrap.hidden = true; return; }
  wrap.hidden = false;
  const on = !!c.dry_run;
  wrap.className = "dryrun-control " + (on ? "on" : "off");
  $("dryrun-toggle").checked = on;
  $("dryrun-state").textContent = on ? "ON — nothing reaches the broker" : "OFF — LIVE";
  wrap.title = c.source === "env"
    ? "Following DASHBOARD_DRY_RUN_DEFAULT — no operator override set yet"
    : "Operator override" + (c.updated_ts ? ", set " + c.updated_ts.slice(0, 19).replace("T", " ") : "");
}

async function postDryRun(enabled, confirm) {
  dryRunApplying = true;
  $("dryrun-control").classList.add("busy");
  try {
    const body = confirm ? { enabled: enabled, confirm: true } : { enabled: enabled };
    const res = await fetch(withToken("/api/dryrun"), {
      method: "POST", headers: authHeaders({ "Content-Type": "application/json" }),
      body: JSON.stringify(body),
    });
    if (!res.ok) {
      const err = await res.json().catch(() => ({}));
      alert(err.error || ("could not change dry-run (HTTP " + res.status + ")"));
    }
  } catch (e) {
    alert("could not change dry-run: " + e);
  } finally {
    dryRunApplying = false;
    $("dryrun-control").classList.remove("busy");
    await refreshPipeline();
  }
}

$("dryrun-toggle").addEventListener("change", (e) => {
  if (dryRunApplying) return;
  if (e.target.checked) { postDryRun(true, false); return; }
  // Going live: revert the visual and let the modal decide.
  e.target.checked = true;
  $("dryrun-scrim").style.display = "flex";
});
$("dryrun-cancel").addEventListener("click", () => { $("dryrun-scrim").style.display = "none"; });
$("dryrun-scrim").addEventListener("click", (e) => { if (e.target.id === "dryrun-scrim") $("dryrun-scrim").style.display = "none"; });
$("dryrun-confirm").addEventListener("click", () => {
  $("dryrun-scrim").style.display = "none";
  postDryRun(false, true);
});

let pendingResubmit = null; // fields for the webhook-test send, or null

function renderResubmitModal(d) {
  const n = (d.commands || []).length;
  $("resubmit-title").textContent = d.ok ? `Place ${n} order${n === 1 ? "" : "s"}?` : "Cannot place order";
  const warnEls = (d.warnings || []).map(w => `<div class="neg" style="margin:.2rem 0">${esc(w)}</div>`).join("");
  const errEls = (d.errors || []).map(w => `<div class="neg" style="margin:.2rem 0">${esc(w)}</div>`).join("");
  $("resubmit-warnings").innerHTML = warnEls + errEls;
  $("resubmit-commands").textContent = n ? d.commands.join("\\n") : "(no commands -- see errors above)";
  $("resubmit-results").innerHTML = "";
  const btn = $("resubmit-confirm");
  btn.disabled = !d.ok;
  btn.textContent = d.ok ? `Place these ${n} order${n === 1 ? "" : "s"}` : "Place orders";
  $("resubmit-scrim").style.display = "flex";
}

function gatherWebhookTestFields() {
  const command = $("wt-command").value;
  return {
    command,
    symbol: $("wt-symbol").value,
    vol_lots: $("wt-vol-lots").value,
    risk_mode: $("wt-risk-mode").value,
    risk: $("wt-risk").value,
    sl: $("wt-sl").value,
    tp: $("wt-tp").value,
    comment: $("wt-comment").value,
  };
}

async function openWebhookTestPreview() {
  const fields = gatherWebhookTestFields();
  pendingResubmit = fields;
  try {
    const res = await fetch(withToken("/api/webhook-test/send"), {
      method: "POST", headers: authHeaders({ "Content-Type": "application/json" }), body: JSON.stringify(fields),
    });
    renderResubmitModal(await res.json());
  } catch (e) {
    renderResubmitModal({ ok: false, commands: [], warnings: [], errors: ["preview request failed: " + e] });
  }
}

async function confirmResubmit() {
  if (!pendingResubmit) return;
  const body = Object.assign({ confirm: true }, pendingResubmit);
  const btn = $("resubmit-confirm");
  btn.disabled = true;
  try {
    const res = await fetch(withToken("/api/webhook-test/send"), {
      method: "POST", headers: authHeaders({ "Content-Type": "application/json" }), body: JSON.stringify(body),
    });
    const d = await res.json();
    if (res.ok) {
      $("resubmit-results").innerHTML = (d.results || []).map(r =>
        `<div>${r.http_status === 200 ? '<span class="badge ok">sent</span>' : '<span class="badge bad">failed</span>'} trace ${esc(r.trace_id || "-")} (HTTP ${r.http_status})</div>`
      ).join("") || `<div class="${d.ok ? "pos" : "neg"}">${d.ok ? "sent" : "not sent"}</div>`;
    } else {
      $("resubmit-results").innerHTML = `<div class="neg">${esc(d.error || ("HTTP " + res.status))}</div>`;
      btn.disabled = false;
    }
  } catch (e) {
    $("resubmit-results").innerHTML = `<div class="neg">resubmit request failed: ${esc(String(e))}</div>`;
    btn.disabled = false;
  }
}

document.addEventListener("click", (e) => {
  const webhookTestBtn = e.target.closest("#webhook-test-preview-btn");
  if (webhookTestBtn) { openWebhookTestPreview(); return; }
  const rcancel = e.target.closest("#resubmit-cancel");
  if (rcancel) { $("resubmit-scrim").style.display = "none"; pendingResubmit = null; return; }
});
$("resubmit-confirm").addEventListener("click", confirmResubmit);
$("resubmit-scrim").addEventListener("click", e => { if (e.target.id === "resubmit-scrim") { $("resubmit-scrim").style.display = "none"; pendingResubmit = null; } });

async function refreshSummary() {
  const res = await fetch(withToken("/api/summary?days=" + currentDays + sourceQS()), { headers: authHeaders() });
  if (!res.ok) { $("meta").textContent = "refresh failed: " + res.status + " " + (await res.text()); return; }
  const s = await res.json();
  lastSummary = s;
  const bs = s.orders.by_source, tv = bs.tradingview;
  const c = s.mt5.available ? s.mt5.closed : {count:0,wins:0,losses:0,net:0,buys:0,sells:0,rows:[],by_source:{},daily:[],review_queue:[]};
  const winRate = c.count ? Math.round(100 * c.wins / c.count) : null;
  const timeLabel = s.mt5.available ? (s.mt5.time_label || "broker time") : "UTC";
  const risk = s.risk || {};
  const eqc = risk.equity_curve || { points: [], estimated: true, max_drawdown_pct: 0, current_drawdown_pct: 0 };
  const comp = risk.compliance || {};
  const overCap = comp.max_single_order_risk != null && comp.cap_usd && comp.max_single_order_risk > comp.cap_usd;
  const capStatus = comp.cap_usd ? (overCap ? "Breached" : "OK") : "—";

  $("kpi-label").textContent = `Net P/L · ${currentDays} days` + (currentSource ? ` · ${sourceLabel(currentSource)}` : "");
  $("kpi-net").textContent = money(c.net);
  $("kpi-net").className = "kpi-value number " + cls(c.net);
  $("acct").innerHTML = s.mt5.available
    ? `<b>MT5 demo ${s.mt5.account}</b><br>balance $${s.mt5.balance.toFixed(2)} · equity $${s.mt5.equity.toFixed(2)}`
    : '<span class="neg">MT5 offline</span>';

  $("cards").innerHTML =
    card(`Net P/L · ${currentDays}d`, money(c.net), cls(c.net), `${c.wins}W / ${c.losses}L`) +
    card("Win rate", winRate === null ? "—" : winRate + "%", winRate === null ? "" : (winRate >= 50 ? "pos" : "neg"), c.count ? `${c.count} closed` : "no trades yet") +
    card(`Trades closed · ${currentDays}d`, c.count, "", `${c.buys||0} buy · ${c.sells||0} sell`) +
    card("Open positions", s.mt5.available ? s.mt5.open.length : "—", "", s.mt5.available ? `floating ${money(s.mt5.open_floating || 0)}` : "MT5 offline") +
    card("Equity", s.mt5.available ? "$" + s.mt5.equity.toFixed(2) : "—", "", s.mt5.available ? `balance $${s.mt5.balance.toFixed(2)} · account-wide` : "") +
    card("Max drawdown", eqc.max_drawdown_pct.toFixed(2) + "%", eqc.max_drawdown_pct > 0 ? "neg" : "", `current ${eqc.current_drawdown_pct.toFixed(2)}% · account-wide`) +
    card("Risk cap status", capStatus, capStatus === "Breached" ? "neg" : (capStatus === "OK" ? "pos" : ""), comp.cap_usd ? `$${comp.cap_usd.toFixed(2)} cap` : "no cap set");

  $("chip-orders").textContent = `${tv.total} total`;
  $("chip-journal").textContent = `${s.journal.entries} entries`;
  $("chip-srcpl").textContent = `${currentDays}d window`;

  table("tbl-orders", ["time (UTC)","source","command","symbol","size","SL","TP","status"], s.orders.recent.map(r =>
    `<tr><td class="muted number">${esc((r.ts||"").slice(5,19).replace("T"," "))}</td><td>${srcBadge(r.source)}</td>
     <td>${sideBadge(_sideOf(r.command))} <span class="muted">${esc(r.command)}</span></td><td>${esc(r.symbol)}</td>
     <td class="number">${r.risk ? "risk $"+esc(r.risk) : esc(r.volume)}</td>
     <td class="number">${esc(r.sl)}</td><td class="number">${esc(r.tp)}</td>
     <td><span class="badge ${r.status==="rejected"?"bad":"ok"}">${esc(r.status)}</span>${r.error?` <span class="neg" title="${esc(r.error)}">!</span>`:""}</td></tr>`));

  const openRows = s.mt5.open || [];
  const floatTotal = s.mt5.available ? (s.mt5.open_floating || 0) : 0;
  $("chip-open").textContent = `${openRows.length} open`;
  table("tbl-open", ["ticket","source","symbol","side","lot","entry","SL","TP","floating P/L"], openRows.map(p =>
    `<tr><td class="muted number">${p.ticket}</td><td>${srcBadge(p.source)}</td><td>${esc(p.symbol)}</td><td>${sideBadge(p.side)}</td>
     <td class="number">${p.volume}</td><td class="number">${p.entry}</td><td class="number">${p.sl}</td><td class="number">${p.tp}</td>
     <td class="number ${cls(p.profit)}">${money(p.profit)}</td></tr>`),
    openRows.length ? `<tr><td colspan="8" class="muted">Total floating</td><td class="number ${cls(floatTotal)}">${money(floatTotal)}</td></tr>` : "");

  const bySrc = c.by_source || {};
  $("srcpl").innerHTML =
    card("TradingView", money(bySrc.tradingview || 0), cls(bySrc.tradingview || 0), `net over ${currentDays}d`, currentSource === "tradingview" ? "highlight" : "") +
    card("Other EA / manual", money(bySrc.other || 0), cls(bySrc.other || 0), `net over ${currentDays}d`, currentSource === "other" ? "highlight" : "");

  renderChart(c.daily || []);
  renderCumulative(c.daily || []);
  renderSymbolChart(c.rows || []);
  renderDonutWinLoss(c.wins || 0, c.losses || 0);
  renderDonutBuySell(c.buys || 0, c.sells || 0);
  renderRisk(s.risk);

  const rq = (c.review_queue || []).slice(0, 8);
  if (rq.length) {
    $("review-strip").style.display = "flex";
    $("review-strip").innerHTML = rq.map(r =>
      `<div class="review-chip">Unjournaled: ${r.side.toUpperCase()} ${esc(r.symbol)} ${money(r.profit)}
       <button type="button" data-ticket="${esc(r.ticket)}">Journal</button></div>`).join("");
  } else {
    $("review-strip").style.display = "none";
    $("review-strip").innerHTML = "";
  }

  const closedHeader = [`closed (${timeLabel})`,"source","symbol","side","lot","entry","close","P/L","setup","emotion","rating","",""];
  table("tbl-closed", closedHeader, c.rows.map(r => {
    const j = r.journal || {};
    return `<tr><td class="muted number">${esc((r.time||"").slice(5,19).replace("T"," "))}</td><td>${srcBadge(r.source)}</td>
     <td>${esc(r.symbol)}</td><td>${sideBadge(r.side)}</td><td class="number">${r.volume}</td>
     <td class="number">${r.entry ?? ""}</td><td class="number">${r.close}</td>
     <td class="number ${cls(r.profit)}">${money(r.profit)}</td>
     <td>${esc(j.setup||"")}</td><td class="muted">${esc(j.emotion||"")}</td>
     <td class="stars">${j.rating ? "\u2605".repeat(j.rating) : ""}</td>
     <td>${j.reviewed ? '<span class="badge ok">reviewed</span>' : ""}</td>
     <td><button class="jbtn" type="button" data-ticket="${esc(r.ticket)}">Journal</button></td></tr>`;
  }));

  $("export-trades").href = withToken("/api/export/trades.csv?days=" + currentDays + sourceQS());
  $("export-journal").href = withToken("/api/export/journal.csv");
  $("export-weekly-xlsx").href = withToken("/api/export/weekly.xlsx?days=7" + sourceQS());

  $("meta").textContent = "updated " + s.updated + " — auto-refreshes every 10s — journal entries are stored locally and mirror the ReyLens schema";
}

function renderScorecard(sc) {
  $("chip-scorecard").textContent = `${sc.days}d window`;
  table("tbl-scorecard",
    ["channel","received","posted","rejected","orders executed","open/pending","wins","losses","net P/L","avg R"],
    (sc.rows || []).map(r =>
      `<tr><td>${esc(r.channel)}</td><td class="number">${r.received}</td><td class="number">${r.posted}</td>
       <td class="number">${r.rejected}</td><td class="number">${r.orders_executed}</td>
       <td class="number muted">${r.orders_open_pending}</td>
       <td class="number pos">${r.wins}</td><td class="number neg">${r.losses}</td>
       <td class="number ${cls(r.net_pl)}">${money(r.net_pl)}</td>
       <td class="number">${(r.avg_r === null || r.avg_r === undefined) ? "—" : r.avg_r.toFixed(2)}</td></tr>`));
}

async function refreshScorecard() {
  // Deliberately unfiltered -- the scorecard IS the source comparison.
  const res = await fetch(withToken("/api/scorecard?days=" + currentDays), { headers: authHeaders() });
  if (!res.ok) return;
  renderScorecard(await res.json());
}

function renderRisk(risk) {
  if (!risk) return;
  const live = risk.live || {};
  const eqc = risk.equity_curve || { points: [], estimated: true, max_drawdown_pct: 0, current_drawdown_pct: 0 };
  const comp = risk.compliance || {};

  // Equity curve + margin are account-level and always unfiltered.
  $("chip-equity").textContent = "account-wide" + (eqc.estimated ? " · estimated" : "");
  $("chip-risk").textContent = `${risk.days}d window` +
    (currentSource ? ` · compliance filtered: ${sourceLabel(currentSource)}` : " · stack-level (tradingview)");

  $("risk-cards").innerHTML =
    card("Margin used", live.available ? "$" + live.margin.toFixed(2) : "—", "", live.available ? `free $${live.margin_free.toFixed(2)} · account-wide` : "MT5 offline") +
    card("Margin level", (live.available && live.margin_level != null) ? live.margin_level.toFixed(1) + "%" : "—", "", "account-wide") +
    card("Open lots", live.available ? live.open_lots : "—", "", "account-wide") +
    card("Current drawdown", eqc.current_drawdown_pct.toFixed(2) + "%", eqc.current_drawdown_pct > 0 ? "neg" : "", "account-wide");

  renderEquityCurve(eqc.points || [], eqc.estimated);

  const rejCount = (comp.risk_cap_rejections || {}).count || 0;
  const overCap = comp.max_single_order_risk != null && comp.cap_usd && comp.max_single_order_risk > comp.cap_usd;
  $("compliance-strip").innerHTML =
    card("Risk cap", comp.cap_usd ? "$" + comp.cap_usd.toFixed(2) : "—", "", "EA_SHIM_RISK_USD") +
    card("Risk-sized orders", (comp.risk_sized_orders || {}).count ?? 0) +
    card("Risk-cap rejections", rejCount, rejCount ? "neg" : "") +
    card("Max single-order risk", comp.max_single_order_risk != null ? "$" + comp.max_single_order_risk.toFixed(2) : "—",
      overCap ? "neg" : "", comp.max_vs_cap_pct != null ? `${comp.max_vs_cap_pct}% of cap` : "");

  const rej = (comp.risk_cap_rejections || {}).rows || [];
  table("tbl-risk-rejections", ["time (UTC)","symbol","trace id","rejection reason"], rej.map(r =>
    `<tr><td class="muted number">${esc((r.ts||"").slice(5,19).replace("T"," "))}</td><td>${esc(r.symbol)}</td>
     <td class="muted">${esc((r.trace_id||"").slice(0,10))}…</td><td class="neg">${esc(r.error)}</td></tr>`));
}

function renderEquityCurve(points, estimated) {
  const el = $("chart-equity");
  if (!points.length) { el.innerHTML = '<span class="muted">no data in this window</span>'; return; }
  const w = 1000, h = 120, padB = 18, padT = 10, padL = 2, padR = 2;
  const vals = points.map(p => p.equity);
  const min = Math.min(...vals), max = Math.max(...vals);
  const range = (max - min) || 1;
  const n = points.length;
  const stepX = n > 1 ? (w - padL - padR) / (n - 1) : 0;
  const yOf = v => padT + (h - padT - padB) * (1 - (v - min) / range);
  let path = "", markers = "";
  points.forEach((p, i) => {
    const x = padL + i * stepX;
    const y = yOf(p.equity);
    path += (i === 0 ? "M" : "L") + x.toFixed(1) + "," + y.toFixed(1) + " ";
    markers += `<circle cx="${x.toFixed(1)}" cy="${y.toFixed(1)}" r="7" fill="transparent"><title>${esc((p.ts||"").slice(0,10))}: ${money(p.equity)}</title></circle>`;
  });
  const area = path + `L${(padL + (n - 1) * stepX).toFixed(1)},${(h - padB).toFixed(1)} L${padL},${(h - padB).toFixed(1)} Z`;
  const tickEvery = Math.max(1, Math.ceil(n / 8));
  let labels = "";
  points.forEach((p, i) => {
    if (i % tickEvery === 0 || i === n - 1) {
      const x = padL + i * stepX;
      labels += `<text x="${x.toFixed(1)}" y="${h-4}" font-size="10" text-anchor="middle" fill="var(--color-text-muted)">${esc((p.ts||"").slice(5,10))}</text>`;
    }
  });
  el.innerHTML = `<svg viewBox="0 0 ${w} ${h}" preserveAspectRatio="none" role="img" aria-label="Equity curve${estimated ? ' (estimated)' : ''}">
    <path d="${area}" fill="var(--color-primary-glow)" stroke="none"/>
    <path d="${path}" fill="none" stroke="var(--color-primary)" stroke-width="2"/>
    ${markers}${labels}
  </svg>` + (estimated
    ? '<div class="muted" style="font-size:0.68rem;margin-top:0.35rem">estimated: balance now minus reverse-cumulated daily closed P/L (fewer than 2 live equity snapshots in this window)</div>'
    : '');
}

function renderCumulative(daily) {
  const el = $("chart-cum");
  if (!daily.length) { el.innerHTML = '<span class="muted">no data in this window</span>'; $("chip-cum").textContent = ""; return; }
  let running = 0;
  const points = daily.map(d => { running += d.pnl; return { date: d.date, cum: running }; });
  const w = 1000, h = 120, padB = 18, padT = 10, padL = 2, padR = 2;
  const vals = points.map(p => p.cum).concat([0]);
  const min = Math.min(...vals), max = Math.max(...vals);
  const range = (max - min) || 1;
  const n = points.length;
  const stepX = n > 1 ? (w - padL - padR) / (n - 1) : 0;
  const yOf = v => padT + (h - padT - padB) * (1 - (v - min) / range);
  const zeroY = yOf(0);
  let path = "", markers = "";
  points.forEach((p, i) => {
    const x = padL + i * stepX;
    const y = yOf(p.cum);
    path += (i === 0 ? "M" : "L") + x.toFixed(1) + "," + y.toFixed(1) + " ";
    markers += `<circle cx="${x.toFixed(1)}" cy="${y.toFixed(1)}" r="7" fill="transparent"><title>${esc(p.date)}: ${money(p.cum)}</title></circle>`;
  });
  const last = points[points.length - 1].cum;
  const color = last >= 0 ? "var(--color-profit)" : "var(--color-loss)";
  const glow = last >= 0 ? "var(--color-profit-dim)" : "var(--color-loss-dim)";
  const area = path + `L${(padL + (n - 1) * stepX).toFixed(1)},${zeroY.toFixed(1)} L${padL},${zeroY.toFixed(1)} Z`;
  const tickEvery = Math.max(1, Math.ceil(n / 8));
  let labels = "";
  points.forEach((p, i) => {
    if (i % tickEvery === 0 || i === n - 1) {
      const x = padL + i * stepX;
      labels += `<text x="${x.toFixed(1)}" y="${h-4}" font-size="10" text-anchor="middle" fill="var(--color-text-muted)">${esc(p.date.slice(5))}</text>`;
    }
  });
  el.innerHTML = `<svg viewBox="0 0 ${w} ${h}" preserveAspectRatio="none" role="img" aria-label="Cumulative net P/L">
    <line x1="0" y1="${zeroY.toFixed(1)}" x2="${w}" y2="${zeroY.toFixed(1)}" stroke="var(--color-border-light)" stroke-width="1"/>
    <path d="${area}" fill="${glow}" stroke="none"/>
    <path d="${path}" fill="none" stroke="${color}" stroke-width="2"/>
    ${markers}${labels}
  </svg>`;
  $("chip-cum").textContent = money(last) + ` over ${currentDays}d` + (currentSource ? ` · ${sourceLabel(currentSource)}` : "");
}

function renderSymbolChart(rows) {
  const el = $("chart-symbol");
  const bySymbol = {};
  (rows || []).forEach(r => { bySymbol[r.symbol] = (bySymbol[r.symbol] || 0) + r.profit; });
  let entries = Object.entries(bySymbol).map(([symbol, net]) => ({ symbol, net: Math.round(net * 100) / 100 }));
  entries.sort((a, b) => Math.abs(b.net) - Math.abs(a.net));
  const total = entries.length;
  entries = entries.slice(0, 8);
  if (!entries.length) { el.innerHTML = '<span class="muted">no data in this window</span>'; $("chip-symbol").textContent = ""; return; }
  $("chip-symbol").textContent = `top ${entries.length} of ${total}`;
  const w = 1000, rowH = 28, padL = 92, padR = 74, padT = 6, padB = 6;
  const n = entries.length;
  const h = padT + padB + n * rowH;
  const max = Math.max(1, ...entries.map(e => Math.abs(e.net)));
  const scale = (w - padL - padR) / max;
  let bars = "";
  entries.forEach((e, i) => {
    const y = padT + i * rowH;
    const barW = Math.max(Math.abs(e.net) * scale, e.net !== 0 ? 1 : 0);
    const x = e.net >= 0 ? padL : padL - barW;
    const color = e.net >= 0 ? "var(--color-profit)" : "var(--color-loss)";
    const midY = (y + rowH / 2 + 3.5).toFixed(1);
    bars += `<text x="${padL - 8}" y="${midY}" font-size="10" text-anchor="end" fill="var(--color-text)">${esc(e.symbol)}</text>`;
    bars += `<rect x="${x.toFixed(1)}" y="${(y + 5).toFixed(1)}" width="${barW.toFixed(1)}" height="${(rowH - 10).toFixed(1)}" fill="${color}"><title>${esc(e.symbol)}: ${money(e.net)}</title></rect>`;
    const labelX = e.net >= 0 ? padL + barW + 6 : padL - barW - 6;
    const anchor = e.net >= 0 ? "start" : "end";
    bars += `<text x="${labelX.toFixed(1)}" y="${midY}" font-size="10" text-anchor="${anchor}" fill="var(--color-text-muted)" class="number">${money(e.net)}</text>`;
  });
  el.innerHTML = `<svg viewBox="0 0 ${w} ${h}" preserveAspectRatio="xMidYMid meet" role="img" aria-label="P/L by symbol" style="height:${Math.max(120, h * 0.24)}px">
    <line x1="${padL}" y1="0" x2="${padL}" y2="${h}" stroke="var(--color-border-light)" stroke-width="1"/>
    ${bars}
  </svg>`;
}

function donutSVG(segments, ariaLabel, centerLabel, centerSub) {
  const total = segments.reduce((a, s) => a + s.value, 0);
  const size = 170, r = 58, cx = size / 2, cy = size / 2, sw = 20;
  const legend = segments.map(s => `<span><span class="dot" style="background:${s.color}"></span>${esc(s.label)}</span>`).join("");
  if (!total) {
    return `<div><svg viewBox="0 0 ${size} ${size}" role="img" aria-label="${esc(ariaLabel)}">
      <circle cx="${cx}" cy="${cy}" r="${r}" fill="none" stroke="var(--color-border)" stroke-width="${sw}"/>
      <text x="${cx}" y="${cy+4}" text-anchor="middle" font-size="11" fill="var(--color-text-muted)">no data</text>
    </svg><div class="muted" style="text-align:center;font-size:0.72rem;margin-top:.4rem">no data in this window</div></div>`;
  }
  const circumference = 2 * Math.PI * r;
  let offset = 0, arcs = "";
  segments.forEach(s => {
    if (!s.value) return;
    const frac = s.value / total;
    const len = frac * circumference;
    arcs += `<circle cx="${cx}" cy="${cy}" r="${r}" fill="none" stroke="${s.color}" stroke-width="${sw}" stroke-dasharray="${len.toFixed(2)} ${(circumference-len).toFixed(2)}" stroke-dashoffset="${(-offset).toFixed(2)}" transform="rotate(-90 ${cx} ${cy})"><title>${esc(s.label)}</title></circle>`;
    offset += len;
  });
  return `<div><svg viewBox="0 0 ${size} ${size}" role="img" aria-label="${esc(ariaLabel)}">
    ${arcs}
    <text x="${cx}" y="${cy-2}" text-anchor="middle" font-size="20" font-weight="600" fill="var(--color-text)">${esc(centerLabel)}</text>
    <text x="${cx}" y="${cy+16}" text-anchor="middle" font-size="10" fill="var(--color-text-muted)">${esc(centerSub||"")}</text>
  </svg><div class="donut-legend">${legend}</div></div>`;
}

function renderDonutWinLoss(wins, losses) {
  const total = wins + losses;
  const rate = total ? Math.round(100 * wins / total) : null;
  $("chart-winloss").innerHTML = donutSVG(
    [{ value: wins, color: "var(--color-profit)", label: `Wins (${wins})` }, { value: losses, color: "var(--color-loss)", label: `Losses (${losses})` }],
    "Win/loss split", rate === null ? "—" : rate + "%", "win rate"
  );
}

function renderDonutBuySell(buys, sells) {
  $("chart-buysell").innerHTML = donutSVG(
    [{ value: buys, color: "var(--color-primary)", label: `Buys (${buys})` }, { value: sells, color: "var(--color-gold)", label: `Sells (${sells})` }],
    "Buy/sell split", String(buys + sells), "orders"
  );
}

let calMonth = null, calStackOnly = true, lastCalendar = null;

function calQuery() {
  return "?stack_only=" + (calStackOnly ? "1" : "0") + (calMonth ? "&month=" + encodeURIComponent(calMonth) : "") + sourceQS();
}

async function refreshCalendar() {
  const res = await fetch(withToken("/api/calendar" + calQuery()), { headers: authHeaders() });
  if (!res.ok) { $("cal-grid").innerHTML = '<span class="muted">calendar failed to load</span>'; return; }
  const data = await res.json();
  calMonth = data.month;
  lastCalendar = data;
  renderCalendar(data);
}

function renderCalendar(data) {
  $("cal-label").textContent = data.month;
  const srcNote = data.source ? sourceLabel(data.source) : (data.stack_only ? "stack only" : "all sources");
  $("chip-calendar").textContent = money(data.total) + " · " + srcNote;
  const days = data.days || [];
  const dow = ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"];
  let html = dow.map(d => `<div class="cal-dow">${d}</div>`).join("");
  const leading = days.length ? days[0].weekday : 0;
  for (let i = 0; i < leading; i++) html += '<div class="cal-cell empty"></div>';
  days.forEach(d => {
    const cellCls = d.count === 0 ? "" : (d.net >= 0 ? "win" : "loss");
    const dayNum = parseInt(d.date.slice(8, 10), 10);
    html += `<div class="cal-cell ${cellCls}"><div class="d">${dayNum}</div>` +
      (d.count ? `<div class="amt ${cls(d.net)}">${money(d.net)}</div><div class="muted" style="font-size:0.62rem">${d.count} trade${d.count===1?"":"s"}</div>` : "") +
      `</div>`;
  });
  $("cal-grid").innerHTML = html;
}

function renderChart(daily) {
  const el = $("chart-daily");
  if (!daily.length) { el.innerHTML = '<span class="muted">no data in this window</span>'; return; }
  const w = 1000, h = 120, padB = 18, padT = 6;
  const max = Math.max(1, ...daily.map(d => Math.abs(d.pnl)));
  const n = daily.length;
  const bw = w / n;
  const zero = padT + (h - padT - padB) / 2;
  const scale = ((h - padT - padB) / 2) / max;
  const tickEvery = Math.max(1, Math.ceil(n / 10));
  let bars = "", labels = "";
  daily.forEach((d, i) => {
    const x = i * bw + bw * 0.15;
    const bwid = bw * 0.7;
    const barH = Math.max(Math.abs(d.pnl) * scale, d.pnl !== 0 ? 1 : 0);
    const y = d.pnl >= 0 ? zero - barH : zero;
    const color = d.pnl >= 0 ? "var(--color-profit)" : "var(--color-loss)";
    bars += `<rect x="${x.toFixed(1)}" y="${y.toFixed(1)}" width="${bwid.toFixed(1)}" height="${barH.toFixed(1)}" fill="${color}"><title>${esc(d.date)}: ${money(d.pnl)}</title></rect>`;
    if (i % tickEvery === 0 || i === n - 1) {
      labels += `<text x="${(x + bwid/2).toFixed(1)}" y="${h - 4}" font-size="10" text-anchor="middle" fill="var(--color-text-muted)">${esc(d.date.slice(5))}</text>`;
    }
  });
  el.innerHTML = `<svg viewBox="0 0 ${w} ${h}" preserveAspectRatio="none" role="img" aria-label="Daily net P/L">
    <line x1="0" y1="${zero.toFixed(1)}" x2="${w}" y2="${zero.toFixed(1)}" stroke="var(--color-border-light)" stroke-width="1"/>
    ${bars}${labels}
  </svg>`;
}

function _sideOf(cmd) {
  cmd = (cmd||"").toLowerCase();
  if (cmd === "closeshortopenlong" || cmd.startsWith("buy")) return "buy";
  if (cmd === "closelongopenshort" || cmd.startsWith("sell")) return "sell";
  return "close";
}
function openModal(ticket) {
  const row = (lastSummary?.mt5?.closed?.rows || []).find(r => String(r.ticket) === String(ticket));
  const j = row?.journal || {};
  $("modal-title").textContent = `Journal — ${row ? row.side.toUpperCase()+" "+row.symbol+" ("+money(row.profit)+")" : "#"+ticket}`;
  $("j-ticket").value = ticket;
  $("j-setup").value = j.setup || "";
  $("j-emotion").value = j.emotion || "";
  $("j-mistakes").value = j.mistakes || "";
  $("j-notes").value = j.notes || "";
  $("j-reviewed").checked = !!j.reviewed;
  setRating(j.rating || 0);
  $("modal-scrim").style.display = "flex";
}
function closeModal() { $("modal-scrim").style.display = "none"; }
function setRating(n) {
  ratingVal = n;
  $("j-rating").innerHTML = [1,2,3,4,5].map(i =>
    `<span class="${i<=n?"on":""}" data-star="${i}">\u2605</span>`).join("");
}
async function saveJournal() {
  const payload = {
    ticket: $("j-ticket").value, setup: $("j-setup").value, emotion: $("j-emotion").value,
    mistakes: $("j-mistakes").value, rating: ratingVal || null, notes: $("j-notes").value,
    reviewed: $("j-reviewed").checked,
  };
  const res = await fetch(withToken("/api/journal"), { method: "POST", headers: authHeaders({"Content-Type":"application/json"}), body: JSON.stringify(payload) });
  if (res.ok) { closeModal(); refresh(); } else { alert("save failed: " + await res.text()); }
}

document.addEventListener("keydown", e => {
  if (e.key !== "Escape") return;
  closeModal();
  $("resubmit-scrim").style.display = "none"; pendingResubmit = null;
});
$("modal-scrim").addEventListener("click", e => { if (e.target.id === "modal-scrim") closeModal(); });
$("j-cancel").addEventListener("click", closeModal);
$("j-save").addEventListener("click", saveJournal);
document.addEventListener("click", e => {
  const themeBtn = e.target.closest("#theme-toggle");
  if (themeBtn) { toggleTheme(); return; }
  const jbtn = e.target.closest("[data-ticket]");
  if (jbtn) { openModal(jbtn.getAttribute("data-ticket")); return; }
  const star = e.target.closest("[data-star]");
  if (star) { setRating(parseInt(star.getAttribute("data-star"), 10)); return; }
  const segBtn = e.target.closest("#seg-days button");
  if (segBtn) {
    currentDays = parseInt(segBtn.getAttribute("data-days"), 10);
    document.querySelectorAll("#seg-days button").forEach(b => b.classList.toggle("active", b === segBtn));
    persistState();
    refresh();
    return;
  }
  const calPrev = e.target.closest("#cal-prev");
  if (calPrev) { calMonth = lastCalendar ? lastCalendar.prev : null; refreshCalendar(); return; }
  const calNext = e.target.closest("#cal-next");
  if (calNext) { calMonth = lastCalendar ? lastCalendar.next : null; refreshCalendar(); return; }
  const calSrcBtn = e.target.closest("#seg-cal-source button");
  if (calSrcBtn) {
    calStackOnly = calSrcBtn.getAttribute("data-source") === "stack";
    document.querySelectorAll("#seg-cal-source button").forEach(b => b.classList.toggle("active", b === calSrcBtn));
    refreshCalendar();
  }
});
document.addEventListener("change", e => {
  if (e.target.id === "source-filter") {
    currentSource = e.target.value;
    persistState();
    refresh();
  }
  if (e.target.id === "wt-command") {
    const isOpen = e.target.value === "BUY" || e.target.value === "SELL";
    $("wt-open-fields").style.display = isOpen ? "contents" : "none";
  }
  if (e.target.id === "wt-risk-mode") {
    $("wt-risk-custom-field").style.display = e.target.value === "custom" ? "" : "none";
  }
});

const THEME_KEY = "execrelay-dashboard-theme";
function applyTheme(t) {
  document.documentElement.classList.toggle("light", t === "light");
  $("theme-toggle").innerHTML = t === "light" ? "&#9728;" : "&#9789;";
}
function toggleTheme() {
  const cur = document.documentElement.classList.contains("light") ? "light" : "dark";
  const next = cur === "light" ? "dark" : "light";
  localStorage.setItem(THEME_KEY, next);
  applyTheme(next);
}
(function initTheme() {
  let saved = null;
  try { saved = localStorage.getItem(THEME_KEY); } catch (e) {}
  const theme = saved || ((window.matchMedia && matchMedia("(prefers-color-scheme: light)").matches) ? "light" : "dark");
  applyTheme(theme);
})();

loadState();
try {
  const savedPage = localStorage.getItem("execrelay-dashboard-page");
  if (["overview", "ignored", "actioned"].includes(savedPage)) showPage(savedPage);
} catch (e) {}
$("source-filter").value = currentSource;
document.querySelectorAll("#seg-days button").forEach(b => b.classList.toggle("active", parseInt(b.getAttribute("data-days"), 10) === currentDays));
persistState();

fetch(withToken("/api/summary"), { headers: authHeaders() }).then(r => r.json()).then(s => {
  const sel = $("j-emotion");
  s.journal.emotions.forEach(e => { const o = document.createElement("option"); o.value = e; o.textContent = e; sel.appendChild(o); });
});
setRating(0); refresh(); setInterval(refresh, 10000);
</script></body></html>""".replace("__LOGO__", LOGO_SVG)


class Handler(BaseHTTPRequestHandler):
    def _host_ok(self) -> bool:
        host = (self.headers.get("Host") or "").strip().lower()
        return host in ALLOWED_HOSTS

    def _token_ok(self) -> bool:
        if not DASHBOARD_TOKEN:
            return True
        supplied = self.headers.get("X-Dashboard-Token", "")
        if not supplied:
            qs = parse_qs(urlsplit(self.path).query)
            supplied = (qs.get("token") or [""])[0]
        return bool(supplied) and hmac.compare_digest(supplied, DASHBOARD_TOKEN)

    def _reject(self, status: int, msg: str) -> None:
        body = msg.encode()
        self.send_response(status)
        self.send_header("Content-Type", "text/plain")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self) -> None:
        if not self._host_ok():
            self._reject(403, "forbidden: bad host")
            return
        path = urlsplit(self.path).path
        if path in ("/health", "/healthz"):
            self._send(200, b'{"status":"ok"}', "application/json")
            return
        if not self._token_ok():
            self._reject(401, "unauthorized")
            return
        if path == "/api/summary":
            days = _parse_days(self.path)
            source = _parse_source(self.path)
            self._send(200, json.dumps(summary(days, source), default=str).encode(), "application/json")
        elif path == "/api/scorecard":
            # Deliberately source-unfiltered: this endpoint IS the source
            # comparison (see scorecard() docstring / PLAN item 1).
            days = _parse_days(self.path)
            self._send(200, json.dumps(scorecard(days), default=str).encode(), "application/json")
        elif path == "/api/ignored":
            days = _parse_days(self.path)
            source = _parse_source(self.path)
            self._send(200, json.dumps(ignored_messages(days, source), default=str).encode(), "application/json")
        elif path == "/api/actioned":
            days = _parse_days(self.path)
            source = _parse_source(self.path)
            self._send(200, json.dumps(actioned_messages(days, source), default=str).encode(), "application/json")
        elif path == "/api/risk":
            days = _parse_days(self.path)
            source = _parse_source(self.path)
            self._send(200, json.dumps(risk_panel(days, source), default=str).encode(), "application/json")
        elif path == "/api/calendar":
            qs = parse_qs(urlsplit(self.path).query)
            month = (qs.get("month") or [""])[0]
            stack_only = (qs.get("stack_only") or ["1"])[0].lower() not in ("0", "false", "")
            source = _parse_source(self.path)
            self._send(
                200, json.dumps(calendar_month(month, stack_only, source), default=str).encode(), "application/json"
            )
        elif path == "/api/export/trades.csv":
            days = _parse_days(self.path)
            source = _parse_source(self.path)
            self._send_csv(trades_csv(days, source), "trades.csv")
        elif path == "/api/export/journal.csv":
            self._send_csv(journal_csv(), "journal.csv")
        elif path == "/api/pipeline":
            self._send(200, json.dumps(pipeline_status(), default=str).encode(), "application/json")
        elif path == "/api/export/weekly.xlsx":
            days = _parse_days(self.path)
            source = _parse_source(self.path)
            xbytes = build_weekly_xlsx(days, source)
            if xbytes is None:
                self._send(
                    501,
                    b"openpyxl is not installed on the server; run: "
                    b"python -m pip install openpyxl",
                    "text/plain",
                )
            else:
                self._send_file(
                    xbytes,
                    "weekly-report.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
        elif path == "/assets/favicon.png":
            try:
                self._send(200, (ASSETS / "favicon.png").read_bytes(), "image/png")
            except OSError:
                self.send_response(404)
                self.end_headers()
        elif path == "/":
            self._send(200, PAGE.encode(), "text/html; charset=utf-8")
        else:
            self.send_response(404)
            self.end_headers()

    _JSON_POST_ROUTES = (
        "/api/journal", "/api/dryrun", "/api/webhook-test/send",
    )

    def do_POST(self) -> None:
        if not self._host_ok():
            self._reject(403, "forbidden: bad host")
            return
        path = urlsplit(self.path).path
        if path not in self._JSON_POST_ROUTES:
            self.send_response(404)
            self.end_headers()
            return
        if not self._token_ok():
            self._reject(401, "unauthorized")
            return
        ctype = (self.headers.get("Content-Type") or "").split(";")[0].strip().lower()
        if ctype != "application/json":
            self._reject(403, "forbidden: Content-Type must be application/json")
            return
        try:
            length = int(self.headers.get("Content-Length", 0))
            payload = json.loads(self.rfile.read(length) or b"{}")
            if path == "/api/journal":
                entry = save_journal_entry(payload)
                self._send(200, json.dumps(entry).encode(), "application/json")
            elif path == "/api/dryrun":
                status, body = set_dry_run_mode(payload)
                self._send(status, json.dumps(body, default=str).encode(), "application/json")
            elif path == "/api/webhook-test/send":
                status, body = send_test_signal(payload)
                self._send(status, json.dumps(body, default=str).encode(), "application/json")
        except (ValueError, json.JSONDecodeError) as exc:
            self._send(400, str(exc).encode(), "text/plain")

    def _send(self, status: int, body: bytes, ctype: str) -> None:
        self.send_response(status)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _send_csv(self, body: bytes, filename: str) -> None:
        self.send_response(200)
        self.send_header("Content-Type", "text/csv; charset=utf-8")
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _send_file(self, body: bytes, filename: str, ctype: str) -> None:
        self.send_response(200)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, _fmt: str, *_args: object) -> None:
        pass


def main() -> None:
    host, port = ADDR.rsplit(":", 1)
    server = ThreadingHTTPServer((host, int(port)), Handler)
    print(time.strftime("%H:%M:%S"), f"trade dashboard listening on http://{ADDR}", flush=True)
    server.serve_forever()


def _cli_digest_now() -> None:
    period_days = 7
    if "--period-days" in sys.argv:
        idx = sys.argv.index("--period-days")
        if idx + 1 < len(sys.argv):
            try:
                period_days = int(sys.argv[idx + 1])
            except ValueError:
                pass
    weekly = period_days >= 7
    text = build_digest_text(period_days, weekly)
    print(text)


if __name__ == "__main__":
    if "--digest-now" in sys.argv:
        _cli_digest_now()
    else:
        main()
